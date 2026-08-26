from __future__ import annotations

import argparse
import ast
import json
import queue
import re
import sys
import threading
import time
import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageTk

DETAIL_SHEET = "评测明细"
REQUIRED_COLUMNS = ("图片ID", "期望结果", "系统像素动作")
IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".bmp", ".webp")
RED = "#ff3b30"
BLUE = "#1687ff"
BG = "#202020"
LIST_BATCH = 150
REDRAW_DELAY_MS = 100


@dataclass(frozen=True)
class ReportCase:
    report_row: int
    source_row: int
    instruction: str
    image_id: str
    expected_raw: str
    system_pixel_raw: str
    vla_correct: bool | None
    comparison: str
    error: str


def log(stage: str, message: str = "") -> None:
    now = time.strftime("%H:%M:%S")
    print(f"[{now}] [{stage}] {message}", flush=True)


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_optional_bool(value: Any) -> bool | None:
    if value is None or cell_text(value) == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in {0, 1}:
        return bool(value)
    text = cell_text(value).casefold()
    if text in {"true", "1", "yes", "y", "正确", "通过", "pass"}:
        return True
    if text in {"false", "0", "no", "n", "错误", "失败", "fail"}:
        return False
    return None


def vla_result_text(value: bool | None) -> str:
    if value is True:
        return "正确"
    if value is False:
        return "错误"
    return "未执行"


def parse_expected_action(raw: str) -> dict[str, Any] | None:
    raw = raw.strip()
    if not raw:
        return None

    value: Any = None
    try:
        value = json.loads(raw)
    except Exception:
        try:
            value = ast.literal_eval(raw)
        except Exception:
            pass

    if isinstance(value, dict):
        if "action" in value or "action_id" in value:
            return value
        for key in ("expected", "expected_result", "target"):
            nested = value.get(key)
            if isinstance(nested, dict):
                return nested
        if any(k in value for k in ("bbox", "box", "coordinate")):
            result = dict(value)
            result.setdefault("action", "click")
            if "box" in result and "bbox" not in result:
                result["bbox"] = result["box"]
            return result

    if isinstance(value, (list, tuple)):
        if len(value) == 4 and all(isinstance(x, (int, float)) for x in value):
            return {"action": "click", "bbox": list(value)}
        if len(value) == 2 and all(isinstance(x, (int, float)) for x in value):
            return {"action": "click", "coordinate": list(value)}

    m = re.search(
        r"(?:bbox|box|bounds)?\s*[:=]?\s*"
        r"[\[\(]\s*(-?\d+(?:\.\d+)?)\s*,\s*"
        r"(-?\d+(?:\.\d+)?)\s*,\s*"
        r"(-?\d+(?:\.\d+)?)\s*,\s*"
        r"(-?\d+(?:\.\d+)?)\s*[\]\)]",
        raw,
        re.I,
    )
    if m:
        return {"action": "click", "bbox": [float(m.group(i)) for i in range(1, 5)]}

    m = re.search(
        r"(?:coordinate|point|click)\s*[:=]?\s*"
        r"[\[\(]\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*[\]\)]",
        raw,
        re.I,
    )
    if m:
        return {"action": "click", "coordinate": [float(m.group(1)), float(m.group(2))]}

    raise ValueError(f"无法解析期望结果：{raw}")


def action_name(action: dict[str, Any] | None) -> str:
    if not action:
        return ""
    return str(action.get("action") or action.get("action_id") or "click")


def parse_system_actions(raw: str) -> list[tuple[str, dict[str, Any]]]:
    """
    解析“系统像素动作”。

    支持：
    1. 单动作：
       {"action":"click","coordinate":[100,200]}

    2. 多 engine：
       {
         "ocr":{"action":"click","coordinate":[100,200]},
         "vla":{"action":"click","bbox":[[10,20,30,40]]}
       }

    返回：
        [("", action)]
    或：
        [("ocr", action1), ("vla", action2)]
    """
    raw = raw.strip()
    if not raw:
        return []

    value: Any = None
    try:
        value = json.loads(raw)
    except Exception:
        try:
            value = ast.literal_eval(raw)
        except Exception as exc:
            raise ValueError(f"无法解析系统像素动作：{raw}") from exc

    if not isinstance(value, dict):
        raise ValueError("系统像素动作必须是 JSON/dict 对象。")

    if "action" in value or "action_id" in value:
        return [("", value)]

    actions: list[tuple[str, dict[str, Any]]] = []
    for engine, action in value.items():
        if isinstance(action, dict) and (
            "action" in action
            or "action_id" in action
            or any(k in action for k in ("bbox", "box", "coordinate", "start_coordinate"))
        ):
            normalized = dict(action)
            if "action" not in normalized and "action_id" not in normalized:
                if "start_coordinate" in normalized:
                    normalized["action"] = "swipe"
                else:
                    normalized["action"] = "click"
            actions.append((str(engine), normalized))

    if not actions:
        raise ValueError(f"系统像素动作中未找到可绘制动作：{raw}")

    return actions


def resolve_report_image(image_id: str, report_dir: Path, image_root: Path | None) -> Path:
    raw = image_id.strip()
    if not raw:
        raise FileNotFoundError("当前条目的图片ID为空。")

    value = Path(raw).expanduser()
    bases: list[Path] = []
    if value.is_absolute():
        bases.append(value)
    else:
        if image_root is not None:
            bases.append(image_root / value)
        bases.append(report_dir / value)
        bases.append(report_dir / "device_captures" / value)

    for base in bases:
        candidates = [base]
        if not base.suffix:
            candidates += [base.with_suffix(ext) for ext in IMAGE_EXTENSIONS]
        for candidate in candidates:
            if candidate.is_file():
                return candidate.resolve()
    raise FileNotFoundError(f"未找到图片：{image_id}")


def load_report_worker(path: Path, out: queue.Queue) -> None:
    """后台线程：只解析 Excel，不调用任何 Tk API。"""
    try:
        log("EXCEL_LOAD_START", str(path))
        started = time.perf_counter()
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if DETAIL_SHEET not in wb.sheetnames:
                raise ValueError(f"评测报告中没有“{DETAIL_SHEET}”工作表。")
            ws = wb[DETAIL_SHEET]
            rows = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                raise ValueError("评测明细为空。")

            header_values = [cell_text(x) for x in header_row]
            headers = {name: i for i, name in enumerate(header_values) if name}
            missing = [name for name in REQUIRED_COLUMNS if name not in headers]
            if missing:
                raise ValueError("评测明细缺少列：" + "、".join(missing))

            def get(row: tuple[Any, ...], name: str, default: Any = None) -> Any:
                idx = headers.get(name)
                return default if idx is None or idx >= len(row) else row[idx]

            cases: list[ReportCase] = []
            for report_row, values in enumerate(rows, start=2):
                image_id = cell_text(get(values, "图片ID"))
                instruction = cell_text(get(values, "任务指令"))
                if not image_id and not instruction:
                    continue
                source_value = get(values, "源行号", report_row)
                source_row = int(source_value) if isinstance(source_value, (int, float)) else report_row
                cases.append(
                    ReportCase(
                        report_row=report_row,
                        source_row=source_row,
                        instruction=instruction,
                        image_id=image_id,
                        expected_raw=cell_text(get(values, "期望结果")),
                        system_pixel_raw=cell_text(get(values, "系统像素动作")),
                        vla_correct=parse_optional_bool(get(values, "VLA正确")),
                        comparison=cell_text(get(values, "判分说明")),
                        error=cell_text(get(values, "错误")),
                    )
                )
                if len(cases) % 1000 == 0:
                    out.put(("excel_progress", len(cases)))

            if not cases:
                raise ValueError("评测明细中没有可查看的用例。")
        finally:
            wb.close()

        elapsed = time.perf_counter() - started
        log("EXCEL_LOAD_DONE", f"{len(cases)} rows, {elapsed:.2f}s")
        out.put(("excel_done", cases))
    except Exception as e:
        log("EXCEL_LOAD_ERROR", repr(e))
        out.put(("excel_error", e))


def load_image_worker(
    request_id: int,
    index: int,
    case: ReportCase,
    report_dir: Path,
    image_root: Path | None,
    out: queue.Queue,
) -> None:
    """后台线程：解析期望结果、读取图片、转 RGB；不创建 ImageTk.PhotoImage。"""
    try:
        log("IMAGE_LOAD_START", f"index={index}, id={case.image_id}")
        started = time.perf_counter()
        errors: list[str] = []
        try:
            expected = parse_expected_action(case.expected_raw)
        except Exception as e:
            expected = None
            errors.append(str(e))

        try:
            system_actions = parse_system_actions(case.system_pixel_raw)
        except Exception as e:
            system_actions = []
            errors.append(str(e))

        image: Image.Image | None = None
        image_path: Path | None = None
        try:
            image_path = resolve_report_image(case.image_id, report_dir, image_root)
            with Image.open(image_path) as src:
                src.load()
                image = src.convert("RGB")
        except Exception as e:
            errors.append(str(e))

        if case.error:
            errors.append(f"评测记录错误：{case.error}")

        elapsed = time.perf_counter() - started
        log("IMAGE_LOAD_DONE", f"index={index}, {elapsed:.2f}s")
        out.put((
            "image_done",
            request_id,
            index,
            image,
            expected,
            system_actions,
            errors,
            image_path,
        ))
    except Exception as e:
        log("IMAGE_LOAD_ERROR", repr(e))
        out.put(("image_error", request_id, index, e))


def swipe_end(action: dict[str, Any], width: int, height: int) -> tuple[float, float] | None:
    end = action.get("end_coordinate")
    if isinstance(end, (list, tuple)) and len(end) == 2:
        return float(end[0]), float(end[1])
    start = action.get("start_coordinate")
    direction = action.get("direction")
    if not isinstance(start, (list, tuple)) or len(start) != 2:
        return None
    if direction not in {"up", "down", "left", "right"}:
        return None
    fraction = {"short": .15, "medium": .30, "long": .55}.get(action.get("distance"), .30)
    x, y = float(start[0]), float(start[1])
    if direction == "left": x -= width * fraction
    elif direction == "right": x += width * fraction
    elif direction == "up": y -= height * fraction
    else: y += height * fraction
    return max(0, min(width - 1, x)), max(0, min(height - 1, y))


class OverlayCanvas(tk.Canvas):
    def __init__(self, master):
        super().__init__(master, bg=BG, highlightthickness=0)
        self.image: Image.Image | None = None
        self.photo: ImageTk.PhotoImage | None = None
        self.expected: dict[str, Any] | None = None
        self.system_actions: list[tuple[str, dict[str, Any]]] = []
        self.scale = 1.0
        self.ox = self.oy = 0
        self._redraw_job: str | None = None
        self._last_size: tuple[int, int] | None = None
        self.bind("<Configure>", self._on_configure)

    def _on_configure(self, _event=None):
        # 窗口尺寸变化会连续触发很多 Configure；只重绘最后一次。
        if self._redraw_job is not None:
            try:
                self.after_cancel(self._redraw_job)
            except Exception:
                pass
        self._redraw_job = self.after(REDRAW_DELAY_MS, self.redraw)

    def set_message(self, text: str):
        self.image = None
        self.expected = None
        self.system_actions = []
        self.photo = None
        self.delete("all")
        w, h = max(1, self.winfo_width()), max(1, self.winfo_height())
        self.create_text(w / 2, h / 2, text=text, fill="white", font=("TkDefaultFont", 15))

    def configure_case(
        self,
        image: Image.Image | None,
        expected: dict[str, Any] | None,
        system_actions: list[tuple[str, dict[str, Any]]] | None = None,
    ):
        self.image = image
        self.expected = expected
        self.system_actions = list(system_actions or [])
        self.redraw()

    def cp(self, point):
        return self.ox + float(point[0]) * self.scale, self.oy + float(point[1]) * self.scale

    def draw_action(self, action: dict[str, Any] | None, color: str, label: str):
        if not action or self.image is None:
            return

        name = action_name(action).casefold()

        if name == "swipe":
            start = action.get("start_coordinate")
            end = swipe_end(action, self.image.width, self.image.height)
            if isinstance(start, (list, tuple)) and len(start) == 2 and end is not None:
                a, z = self.cp(start), self.cp(end)
                self.create_line(
                    *a, *z,
                    fill=color,
                    width=4,
                    arrow=tk.LAST,
                    arrowshape=(14, 18, 7),
                )
                self.create_text(
                    a[0] + 5,
                    a[1] + 5,
                    text=label,
                    fill=color,
                    anchor="nw",
                    font=("TkDefaultFont", 10, "bold"),
                )
            return

        boxes = action.get("bbox", action.get("box"))

        # bbox: [x1,y1,x2,y2]
        if (
            isinstance(boxes, (list, tuple))
            and len(boxes) == 4
            and all(isinstance(v, (int, float)) for v in boxes)
        ):
            boxes = [boxes]

        # bbox: [[x1,y1,x2,y2], ...]
        if isinstance(boxes, (list, tuple)):
            for box in boxes:
                if isinstance(box, (list, tuple)) and len(box) == 4:
                    a, z = self.cp(box[:2]), self.cp(box[2:])
                    self.create_rectangle(
                        *a, *z,
                        outline=color,
                        width=4,
                    )
                    self.create_text(
                        a[0] + 5,
                        a[1] + 5,
                        text=label,
                        fill=color,
                        anchor="nw",
                        font=("TkDefaultFont", 10, "bold"),
                    )

        coordinate = action.get("coordinate")
        if coordinate is None and all(k in action for k in ("x", "y")):
            coordinate = [action["x"], action["y"]]

        if isinstance(coordinate, (list, tuple)) and len(coordinate) == 2:
            x, y = self.cp(coordinate)
            r = 9
            self.create_oval(
                x-r, y-r, x+r, y+r,
                outline=color,
                width=4,
            )
            self.create_line(
                x-r-5, y, x+r+5, y,
                fill=color,
                width=2,
            )
            self.create_line(
                x, y-r-5, x, y+r+5,
                fill=color,
                width=2,
            )
            self.create_text(
                x + 12,
                y + 12,
                text=label,
                fill=color,
                anchor="nw",
                font=("TkDefaultFont", 10, "bold"),
            )

    def draw_expected(self):
        self.draw_action(self.expected, RED, "期望结果")

    def draw_system_actions(self):
        for engine, action in self.system_actions:
            label = "系统像素动作"
            if engine:
                label += f"-{engine.upper()}"
            self.draw_action(action, BLUE, label)

    def redraw(self):
        self._redraw_job = None
        w, h = max(1, self.winfo_width()), max(1, self.winfo_height())
        if self.image is None:
            self.set_message("请选择右侧条目加载图片")
            return

        size_key = (w, h)
        log("CANVAS_DRAW", f"canvas={w}x{h}, image={self.image.width}x{self.image.height}")
        self.delete("all")
        self.scale = min(w / self.image.width, h / self.image.height)
        shown_size = (
            max(1, round(self.image.width * self.scale)),
            max(1, round(self.image.height * self.scale)),
        )
        # BILINEAR 比 LANCZOS 快很多，查看器足够清晰，也更不容易阻塞 GUI。
        shown = self.image.resize(shown_size, Image.Resampling.BILINEAR)
        self.ox = (w - shown_size[0]) // 2
        self.oy = (h - shown_size[1]) // 2
        self.photo = ImageTk.PhotoImage(shown)
        self.create_image(self.ox, self.oy, image=self.photo, anchor="nw")
        self.draw_expected()
        self.draw_system_actions()

        # 图例
        self.create_rectangle(10, 10, 320, 42, fill=BG, outline="")
        self.create_rectangle(18, 19, 29, 30, fill=RED, outline=RED)
        self.create_text(38, 26, text="期望结果", fill="white", anchor="w")
        self.create_rectangle(121, 19, 132, 30, fill=BLUE, outline=BLUE)
        self.create_text(141, 26, text="系统像素动作", fill="white", anchor="w")

        self._last_size = size_key


class ViewerApp:
    def __init__(self, root: tk.Tk, report_path: Path):
        self.root = root
        self.report_path = report_path.resolve()
        self.report_dir = self.report_path.parent
        self.image_root: Path | None = None
        self.cases: list[ReportCase] = []
        self.index = -1
        self.selecting = False
        self.image_request_id = 0
        # 防止同一条记录因为重复 UI 事件被并发加载多次。
        # key = (index, image_root_string)
        self.loading_image_key = None
        self.q: queue.Queue = queue.Queue()
        self.progress = tk.StringVar(value="正在读取 Excel...")
        self.status = tk.StringVar(value="初始化")
        self.instruction = tk.StringVar(value="未选择")
        self.image_id_var = tk.StringVar(value="未选择")
        self.expected_var = tk.StringVar(value="未选择")
        self.system_pixel_var = tk.StringVar(value="未选择")
        self.vla_result = tk.StringVar(value="未选择")
        self.comparison = tk.StringVar(value="未选择")

        root.title(f"评测结果查看器 - {self.report_path.name}")
        root.geometry("1500x900")
        root.minsize(1000, 650)
        self.build()
        root.bind("<Left>", lambda _e: self.navigate(-1))
        root.bind("<Right>", lambda _e: self.navigate(1))
        self.start_excel_load()
        self.root.after(50, self.poll_queue)

    def build(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        top = ttk.Frame(self.root, padding=8)
        top.grid(row=0, column=0, sticky="ew")
        ttk.Button(top, text="上一条", command=lambda: self.navigate(-1)).pack(side=tk.LEFT)
        ttk.Button(top, text="下一条", command=lambda: self.navigate(1)).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="选择图片目录", command=self.choose_image_root).pack(side=tk.LEFT, padx=(14, 6))
        ttk.Label(top, textvariable=self.progress).pack(side=tk.LEFT, padx=12)
        ttk.Label(top, text="红色：期望结果", foreground=RED).pack(side=tk.RIGHT, padx=8)
        ttk.Label(top, text="蓝色：系统像素动作", foreground=BLUE).pack(side=tk.RIGHT)

        panes = ttk.Panedwindow(self.root, orient=tk.HORIZONTAL)
        panes.grid(row=1, column=0, sticky="nsew", padx=8)
        left = ttk.Frame(panes)
        right = ttk.Frame(panes, padding=(8, 0, 0, 0))
        panes.add(left, weight=3)
        panes.add(right, weight=2)

        left.rowconfigure(0, weight=1)
        left.columnconfigure(0, weight=1)
        self.canvas = OverlayCanvas(left)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.canvas.set_message("正在读取 Excel 条目...")

        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=3)
        right.rowconfigure(1, weight=2)
        cols = ("row", "instruction", "image", "vla")
        self.tree = ttk.Treeview(right, columns=cols, show="headings", selectmode="browse")
        for col, title, width in (
            ("row", "行", 55),
            ("instruction", "任务", 260),
            ("image", "图片ID", 190),
            ("vla", "VLA正确", 80),
        ):
            self.tree.heading(col, text=title)
            self.tree.column(col, width=width, anchor="w")
        scroll = ttk.Scrollbar(right, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        detail = ttk.LabelFrame(right, text="当前用例详情", padding=8)
        detail.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(8, 0))
        detail.columnconfigure(1, weight=1)
        for r, (name, var) in enumerate((
            ("任务", self.instruction),
            ("图片ID", self.image_id_var),
            ("期望结果", self.expected_var),
            ("系统像素动作", self.system_pixel_var),
            ("VLA正确", self.vla_result),
            ("判分说明", self.comparison),
        )):
            ttk.Label(detail, text=name + "：").grid(row=r, column=0, sticky="nw", pady=3)
            ttk.Label(detail, textvariable=var, wraplength=520, justify=tk.LEFT).grid(row=r, column=1, sticky="nw", pady=3)

        ttk.Label(self.root, textvariable=self.status, anchor="w", relief="sunken", padding=(6, 3)).grid(row=2, column=0, sticky="ew", pady=(8, 0))

    def start_excel_load(self):
        threading.Thread(target=load_report_worker, args=(self.report_path, self.q), daemon=True).start()

    def poll_queue(self):
        try:
            while True:
                item = self.q.get_nowait()
                kind = item[0]
                if kind == "excel_progress":
                    self.progress.set(f"已读取 {item[1]} 条...")
                elif kind == "excel_done":
                    self.cases = item[1]
                    self.progress.set(f"共 {len(self.cases)} 条，正在构建列表...")
                    log("LIST_BUILD_START", str(len(self.cases)))
                    self._insert_batch(0)
                elif kind == "excel_error":
                    messagebox.showerror("Excel 读取失败", str(item[1]), parent=self.root)
                    self.status.set(str(item[1]))
                elif kind == "image_done":
                    (
                        _,
                        req_id,
                        index,
                        image,
                        expected,
                        system_actions,
                        errors,
                        image_path,
                    ) = item
                    if req_id != self.image_request_id or index != self.index:
                        continue
                    self.loading_image_key = None
                    self.canvas.configure_case(image, expected, system_actions)
                    self.status.set("；".join(errors) if errors else f"已加载：{image_path}")
                elif kind == "image_error":
                    _, req_id, index, error = item
                    if req_id == self.image_request_id and index == self.index:
                        self.loading_image_key = None
                        self.canvas.set_message("图片加载失败")
                        self.status.set(str(error))
        except queue.Empty:
            pass
        self.root.after(50, self.poll_queue)

    def _insert_batch(self, start: int):
        end = min(start + LIST_BATCH, len(self.cases))
        for i in range(start, end):
            case = self.cases[i]
            self.tree.insert("", tk.END, iid=str(i), values=(
                case.source_row,
                case.instruction,
                case.image_id,
                vla_result_text(case.vla_correct),
            ))
        self.progress.set(f"列表 {end}/{len(self.cases)}")
        if end < len(self.cases):
            self.root.after(1, lambda: self._insert_batch(end))
        else:
            log("LIST_BUILD_DONE", str(len(self.cases)))
            self.progress.set(f"共 {len(self.cases)} 条记录")
            self.status.set("条目加载完成；点击右侧条目后才加载对应图片")
            self.canvas.set_message("请选择右侧条目加载图片")

    def choose_image_root(self):
        selected = filedialog.askdirectory(title="选择图片根目录", parent=self.root)
        if selected:
            self.image_root = Path(selected).resolve()
            self.loading_image_key = None
            self.status.set(f"图片目录：{self.image_root}")
            if self.index >= 0:
                self.load_selected_image(self.index)

    def on_select(self, _event=None):
        if self.selecting:
            return
        selected = self.tree.selection()
        if not selected:
            return

        target = int(selected[0])

        # selection_set() 会再次产生 <<TreeviewSelect>>。
        # 如果还是当前条目，必须直接返回，避免事件递归。
        if target == self.index:
            return

        self.show_case(target)

    def show_case(self, index: int):
        if not (0 <= index < len(self.cases)):
            return
        self.index = index
        case = self.cases[index]
        self.progress.set(f"{index+1}/{len(self.cases)}  Excel 源行：{case.source_row}")
        self.instruction.set(case.instruction or "无")
        self.image_id_var.set(case.image_id or "无")
        self.expected_var.set(case.expected_raw or "无")
        self.system_pixel_var.set(case.system_pixel_raw or "无")
        self.vla_result.set(vla_result_text(case.vla_correct))
        self.comparison.set(case.comparison or "无")

        current_selection = self.tree.selection()
        if current_selection != (str(index),):
            self.selecting = True
            try:
                self.tree.selection_set(str(index))
                self.tree.focus(str(index))
                self.tree.see(str(index))
            finally:
                self.selecting = False

        self.load_selected_image(index)

    def load_selected_image(self, index: int):
        if not (0 <= index < len(self.cases)):
            return

        root_key = str(self.image_root) if self.image_root is not None else ""
        load_key = (index, root_key)

        # 同一条、同一图片目录如果已经在加载，禁止重复创建后台线程。
        if self.loading_image_key == load_key:
            log("IMAGE_LOAD_SKIP", f"duplicate index={index}")
            return

        self.loading_image_key = load_key
        self.image_request_id += 1
        req_id = self.image_request_id
        case = self.cases[index]

        self.status.set("正在后台加载当前图片...")
        self.canvas.set_message("正在加载当前图片...")

        threading.Thread(
            target=load_image_worker,
            args=(req_id, index, case, self.report_dir, self.image_root, self.q),
            daemon=True,
        ).start()

    def navigate(self, delta: int):
        if not self.cases:
            return
        if self.index < 0:
            if delta > 0:
                self.show_case(0)
            return
        target = self.index + delta
        if 0 <= target < len(self.cases):
            self.show_case(target)
        else:
            self.status.set("已经是第一条" if target < 0 else "已经是最后一条")


def choose_report(root: tk.Tk) -> Path | None:
    selected = filedialog.askopenfilename(
        title="选择 evaluator 测试报告",
        filetypes=(("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")),
        parent=root,
    )
    return Path(selected) if selected else None


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Async TTK evaluator viewer")
    parser.add_argument("report", nargs="?", type=Path)
    args = parser.parse_args(argv)

    try:
        root = tk.Tk()
    except tk.TclError as e:
        print(f"无法初始化 TTK：{e}", file=sys.stderr)
        return 1

    root.withdraw()
    selected = args.report or choose_report(root)
    if not selected:
        root.destroy()
        return 0

    selected = Path(selected).expanduser().resolve()
    root.deiconify()
    ViewerApp(root, selected)
    root.lift()
    root.after(100, root.focus_force)
    root.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())