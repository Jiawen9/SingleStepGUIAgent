"""TTK editor for reference actions in an evaluation workbook."""
from __future__ import annotations

import argparse, json, os, sys, tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from PIL import Image, ImageTk
from engines.validation import ACTION_SPECS, ActionSpec
from engines.vla.prompts import load_app_prompt

SHEETS = ("测试用例集", "测试用例表")
REQUIRED = ("任务指令", "图片ID", "UI-TREE", "结果输出")
COMMON = frozenset({"click", "swipe", "type", "reject"})
PROJECT_ROOT = Path(__file__).resolve().parent
DEVICE_CAPTURES_DIR = PROJECT_ROOT / "device_captures"

@dataclass(frozen=True)
class Task:
    row: int; instruction: str; image_id: str; xml: str; result: str

def compact(value: dict[str, Any]) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))

def resolve_image_path(image_id: str, base: Path) -> Path:
    path = Path(image_id.strip()).expanduser()
    path = path if path.is_absolute() else base / path
    candidates = [path] + ([] if path.suffix else [path.with_suffix(x) for x in (".png", ".jpg", ".jpeg")])
    for candidate in candidates:
        if candidate.is_file(): return candidate.resolve()
    raise FileNotFoundError(f"未找到图片：{path}")

def actions_for_ui_tree(value: str, base: Path) -> tuple[list[ActionSpec], str]:
    if not value.strip(): return list(ACTION_SPECS), "未提供 UI-TREE；可选择全部动作"
    try:
        if value.lstrip().startswith("<"):
            root = ElementTree.fromstring(value)
        else:
            path = Path(value).expanduser(); path = path if path.is_absolute() else base / path
            root = ElementTree.fromstring(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, ElementTree.ParseError) as error:
        return list(ACTION_SPECS), f"UI-TREE 无法解析；可选择全部动作：{error}"
    packages = []
    for node in root.iter():
        package = (node.get("package") or "").strip()
        if package and package not in packages: packages.append(package)
    for package in packages:
        prompt = load_app_prompt(package)
        if prompt:
            allowed = COMMON | prompt.action_names
            return [x for x in ACTION_SPECS if x.name in allowed], f"已识别包名：{package}（{prompt.app_id}）"
    if packages:
        return (
            [spec for spec in ACTION_SPECS if spec.name in COMMON],
            f"未注册包名：{packages[0]}；仅可选择基础动作",
        )
    return list(ACTION_SPECS), "XML 中没有包名；可选择全部动作"

def swipe_details(start: tuple[int,int], end: tuple[int,int], width: int, height: int) -> tuple[str,str]:
    dx, dy = end[0]-start[0], end[1]-start[1]
    if dx == dy == 0: raise ValueError("滑动起点和终点不能相同。")
    if abs(dx) >= abs(dy): direction, length = ("right" if dx > 0 else "left"), abs(dx)*1000/width
    else: direction, length = ("down" if dy > 0 else "up"), abs(dy)*1000/height
    return direction, ("short" if length <= 200 else "medium" if length <= 400 else "long")

def parse_result(raw: str) -> dict[str,Any] | None:
    if not raw.strip(): return None
    value = json.loads(raw)
    if not isinstance(value, dict) or not isinstance(value.get("action", value.get("action_id")), str):
        raise ValueError("结果输出必须是含 action 或 action_id 的 JSON 对象。")
    action = value.get("action", value.get("action_id"))
    if action not in {spec.name for spec in ACTION_SPECS}:
        raise ValueError(f"未知动作：{action}")
    if action == "click":
        boxes = value.get("bbox")
        if not isinstance(boxes, list) or not boxes or any(
            not isinstance(box, list) or len(box) != 4
            or any(isinstance(number, bool) or not isinstance(number, (int, float)) for number in box)
            for box in boxes
        ):
            raise ValueError("click 的 bbox 必须是非空像素框列表。")
    if action == "swipe":
        start = value.get("start_coordinate")
        if not isinstance(start, list) or len(start) != 2 or any(
            isinstance(number, bool) or not isinstance(number, (int, float)) for number in start
        ):
            raise ValueError("swipe 的 start_coordinate 必须是像素坐标。")
        if value.get("direction") not in {"up", "down", "left", "right"}:
            raise ValueError("swipe 的 direction 无效。")
        if value.get("distance") not in {"short", "medium", "long"}:
            raise ValueError("swipe 的 distance 无效。")
    return value

def make_result(action: str, boxes=None, swipe=None, parameters=None) -> dict[str,Any]:
    if action == "click":
        if not boxes: raise ValueError("click 至少需要一个标注框。")
        return {"action":"click", "bbox":boxes}
    if action == "swipe":
        if not swipe: raise ValueError("请在图片上拖拽标注滑动动作。")
        return {"action":"swipe", **swipe}
    if action == "reject": return {"action_id":"reject", **(parameters or {})}
    return {"action":action, **(parameters or {})}

def select_task_sheet(workbook):
    """Prefer the conventional name, then find a sheet by its required headers."""
    for name in SHEETS:
        if name in workbook.sheetnames:
            sheet = workbook[name]
            headers = {str(cell.value).strip() for cell in sheet[1] if cell.value is not None}
            if set(REQUIRED) <= headers:
                return sheet
    required = set(REQUIRED)
    for worksheet in workbook.worksheets:
        headers = {str(cell.value).strip() for cell in worksheet[1] if cell.value is not None}
        if required <= headers:
            return worksheet
    raise ValueError(
        f"工作簿中没有包含以下列的工作表：{'、'.join(REQUIRED)}"
    )

class DrawCanvas(tk.Canvas):
    def __init__(self, master, changed):
        super().__init__(master, bg="#202020", highlightthickness=0, cursor="crosshair")
        self.changed=changed; self.image=None; self.photo=None; self.mode=""; self.boxes=[]; self.swipe=None; self.swipe_end=None; self.start=None
        self.scale=1.; self.ox=self.oy=0
        self.bind("<Configure>", lambda _e:self.redraw()); self.bind("<Button-1>", self.press); self.bind("<B1-Motion>", self.motion); self.bind("<ButtonRelease-1>", self.release)
        self.bind("<Button-3>", self.delete_box_at)
    def configure_data(self, image, mode="", boxes=None, swipe=None):
        self.image=image; self.mode=mode; self.boxes=[list(x) for x in (boxes or [])]; self.swipe=dict(swipe) if swipe else None; self.swipe_end=None; self.redraw()
    def point(self,e):
        if self.image is None:return None
        x,y=round((e.x-self.ox)/self.scale),round((e.y-self.oy)/self.scale)
        return (x,y) if 0<=x<self.image.width and 0<=y<self.image.height else None
    def cp(self,p): return self.ox+p[0]*self.scale,self.oy+p[1]*self.scale
    def press(self,e):
        if self.mode in {"click","swipe"}: self.start=self.point(e)
    def motion(self,e):
        end=self.point(e)
        if self.start and end:
            self.redraw(); a=self.cp(self.start); b=self.cp(end)
            if self.mode=="click":self.create_rectangle(*a,*b,outline="#00e5ff",width=2)
            else:self.create_line(*a,*b,fill="#ffcc00",width=3,arrow=tk.LAST)
    def release(self,e):
        start,end=self.start,self.point(e); self.start=None
        if not start or not end or start==end:return self.redraw()
        if self.mode=="click":
            self.boxes.append([min(start[0],end[0]),min(start[1],end[1]),max(start[0],end[0]),max(start[1],end[1])])
        else:
            direction,distance=swipe_details(start,end,self.image.width,self.image.height)
            self.swipe={"start_coordinate":[*start],"direction":direction,"distance":distance}; self.swipe_end=end
        self.redraw();self.changed()
    def delete_box_at(self,e):
        if self.mode != "click": return
        point=self.point(e)
        if point is None:return
        for index in range(len(self.boxes)-1,-1,-1):
            x1,y1,x2,y2=self.boxes[index]
            if x1<=point[0]<=x2 and y1<=point[1]<=y2:
                del self.boxes[index];self.redraw();self.changed();return
    def representative_end(self):
        x,y=self.swipe["start_coordinate"]; f={"short":.15,"medium":.3,"long":.55}[self.swipe["distance"]]
        dx,dy=round(self.image.width*f),round(self.image.height*f); d=self.swipe["direction"]
        if d=="left":x-=dx
        elif d=="right":x+=dx
        elif d=="up":y-=dy
        else:y+=dy
        return max(0,min(self.image.width-1,x)),max(0,min(self.image.height-1,y))
    def redraw(self):
        self.delete("all")
        if self.image is None:return self.create_text(max(1,self.winfo_width())/2,max(1,self.winfo_height())/2,text="无法显示图片",fill="white")
        w,h=max(1,self.winfo_width()),max(1,self.winfo_height());self.scale=min(w/self.image.width,h/self.image.height)
        size=(max(1,round(self.image.width*self.scale)),max(1,round(self.image.height*self.scale))); shown=self.image.resize(size,Image.Resampling.LANCZOS)
        self.ox,self.oy=(w-size[0])//2,(h-size[1])//2;self.photo=ImageTk.PhotoImage(shown);self.create_image(self.ox,self.oy,image=self.photo,anchor="nw")
        for i,b in enumerate(self.boxes,1):
            a=self.cp(b[:2]);z=self.cp(b[2:]);self.create_rectangle(*a,*z,outline="#00e5ff",width=2);self.create_text(a[0]+3,a[1]+3,text=str(i),fill="#00e5ff",anchor="nw")
        if self.swipe:
            a=self.cp(self.swipe["start_coordinate"]);z=self.cp(self.swipe_end or self.representative_end());self.create_line(*a,*z,fill="#ffcc00",width=3,arrow=tk.LAST)

class App:
    def __init__(self, root:tk.Tk, path:Path):
        self.root=root;self.path=path.resolve();self.wb=load_workbook(self.path)
        self.ws=select_task_sheet(self.wb);self.headers={str(c.value).strip():c.column for c in self.ws[1] if c.value is not None}
        missing=[x for x in REQUIRED if x not in self.headers]
        if missing:raise ValueError("任务表缺少列："+"、".join(missing))
        self.tasks=[]
        for r in range(2,self.ws.max_row+1):
            vals={x:self.ws.cell(r,self.headers[x]).value for x in REQUIRED}
            if any(v is not None and str(v).strip() for v in vals.values()):self.tasks.append(Task(r,*["" if vals[x] is None else str(vals[x]).strip() for x in ("任务指令","图片ID","UI-TREE","结果输出")]))
        if not self.tasks:raise ValueError("任务表中没有数据行。")
        self.index=0;self.loading=False;self.dirty=False;self.specs=[];self.params={};self.image=None
        self.action=tk.StringVar();self.progress=tk.StringVar();self.package=tk.StringVar();self.status=tk.StringVar();self.selecting=False
        root.title(f"用例参考答案标注 - {self.path.name}");root.geometry("1500x900");root.minsize(1050,680);self.build();root.protocol("WM_DELETE_WINDOW",self.close);self.show(0)
    def build(self):
        self.root.columnconfigure(0,weight=1);self.root.rowconfigure(1,weight=1);top=ttk.Frame(self.root,padding=8);top.grid(row=0,column=0,sticky="ew");top.columnconfigure(9,weight=1)
        for col,(text,cmd) in enumerate((("上一条",lambda:self.nav(-1)),("下一条",lambda:self.nav(1)),("保存当前条",self.save))):ttk.Button(top,text=text,command=cmd).grid(row=0,column=col,padx=(0,5))
        ttk.Label(top,textvariable=self.progress).grid(row=0,column=3,padx=10);ttk.Label(top,text="动作：").grid(row=0,column=4)
        self.combo=ttk.Combobox(top,textvariable=self.action,state="readonly",width=27);self.combo.grid(row=0,column=5,padx=5);self.combo.bind("<<ComboboxSelected>>",self.action_changed)
        self.param_frame=ttk.Frame(top);self.param_frame.grid(row=0,column=6)
        ttk.Button(top,text="撤销标注",command=self.undo).grid(row=0,column=7,padx=(3,3));ttk.Button(top,text="清空标注",command=self.clear).grid(row=0,column=8,padx=(0,8));ttk.Label(top,textvariable=self.package).grid(row=0,column=9,sticky="e")
        panes=ttk.Panedwindow(self.root,orient=tk.HORIZONTAL);panes.grid(row=1,column=0,sticky="nsew",padx=8);left=ttk.Frame(panes);right=ttk.Frame(panes,padding=(8,0,0,0));panes.add(left,weight=3);panes.add(right,weight=2)
        left.rowconfigure(0,weight=1);left.columnconfigure(0,weight=1);self.canvas=DrawCanvas(left,self.canvas_changed);self.canvas.grid(row=0,column=0,sticky="nsew")
        right.columnconfigure(0,weight=1);right.rowconfigure(0,weight=1)
        self.task_list=ttk.Treeview(right,columns=("instruction","action"),show="headings",selectmode="browse")
        self.task_list.heading("instruction",text="任务指令");self.task_list.heading("action",text="动作")
        self.task_list.column("instruction",width=260,anchor="w");self.task_list.column("action",width=300,anchor="w")
        scrollbar=ttk.Scrollbar(right,orient=tk.VERTICAL,command=self.task_list.yview);self.task_list.configure(yscrollcommand=scrollbar.set)
        self.task_list.grid(row=0,column=0,sticky="nsew");scrollbar.grid(row=0,column=1,sticky="ns")
        for index,task in enumerate(self.tasks):
            action_text=self.action_summary(task.result)
            self.task_list.insert("",tk.END,iid=str(index),values=(task.instruction,action_text))
        self.task_list.bind("<<TreeviewSelect>>",self.list_selected)
        ttk.Label(self.root,textvariable=self.status,anchor="w",relief="sunken",padding=(6,3)).grid(row=2,column=0,sticky="ew",pady=(8,0))
    def show(self,index):
        self.loading=True;self.index=index;t=self.tasks[index];self.progress.set(f"{index+1}/{len(self.tasks)}  图片ID：{t.image_id}");self.specs,label=actions_for_ui_tree(t.xml,DEVICE_CAPTURES_DIR);self.package.set(label);self.combo["values"]=[x.name for x in self.specs];self.action.set("");errors=[];self.image=None
        try:
            with Image.open(resolve_image_path(t.image_id,DEVICE_CAPTURES_DIR)) as im:im.load();self.image=im.convert("RGB")
        except OSError as e:errors.append(str(e))
        self.canvas.configure_data(self.image)
        try:value=parse_result(t.result)
        except (ValueError,json.JSONDecodeError) as e:value=None;errors.append(f"已有结果无效，未修改时保持原样：{e}")
        if value is not None:
            existing_action=value.get("action",value.get("action_id"))
            if existing_action not in {spec.name for spec in self.specs}:
                errors.append(f"已有动作 {existing_action} 不属于当前包名，未修改时保持原样")
                value=None
        if value:self.restore(value)
        else:self.build_params();self.set_preview(t.result)
        self.status.set("；".join(errors) if errors else "已加载");self.dirty=False;self.loading=False
        self.selecting=True;self.task_list.selection_set(str(index));self.task_list.focus(str(index));self.task_list.see(str(index));self.selecting=False
    def restore(self,v):
        action=v.get("action",v.get("action_id"))
        self.action.set(action);boxes=[list(map(round,x)) for x in v.get("bbox",[]) if isinstance(x,list) and len(x)==4] if action=="click" else []
        swipe={k:v.get(k) for k in ("start_coordinate","direction","distance")} if action=="swipe" else None;self.canvas.configure_data(self.image,action,boxes,swipe);self.build_params(v);self.refresh_list();self.set_preview(compact(v))
    def build_params(self,existing=None):
        for w in self.param_frame.winfo_children():w.destroy()
        self.params={};spec=next((x for x in self.specs if x.name==self.action.get()),None)
        if not spec or spec.name in {"click","swipe"}:return
        for i,(name,schema) in enumerate(spec.parameters.get("properties",{}).items()):
            ttk.Label(self.param_frame,text=name+"：").grid(row=0,column=i*2);var=tk.StringVar(value="" if not existing else str(existing.get(name,"")));self.params[name]=var
            if "enum" in schema:
                widget=ttk.Combobox(self.param_frame,textvariable=var,values=schema["enum"],state="readonly",width=15)
                if not var.get() and schema["enum"]:var.set(schema["enum"][0])
                widget.bind("<<ComboboxSelected>>",lambda _e:self.mark())
            else:widget=ttk.Entry(self.param_frame,textvariable=var,width=20);widget.bind("<KeyRelease>",lambda _e:self.mark())
            widget.grid(row=0,column=i*2+1,padx=(2,7))
    def action_changed(self,_e=None):
        if self.loading:return
        self.canvas.configure_data(self.image,self.action.get());self.build_params();self.refresh_list();self.mark()
    def result(self):
        if not self.action.get():raise ValueError("请选择动作。")
        params={k:v.get() for k,v in self.params.items()}
        if any(not x.strip() for x in params.values()):raise ValueError("动作参数不能为空。")
        return make_result(self.action.get(),self.canvas.boxes,self.canvas.swipe,params)
    def set_preview(self,text):
        if hasattr(self,"task_list") and self.task_list.exists(str(self.index)):
            values=list(self.task_list.item(str(self.index),"values"));values[1]=self.action_summary(text);self.task_list.item(str(self.index),values=values)
    @staticmethod
    def action_summary(raw):
        try:
            value=json.loads(raw) if isinstance(raw,str) else raw
            if not isinstance(value,dict):return ""
            return compact(value)
        except (ValueError,TypeError):return "未标注" if not raw else "结果格式错误"
    def mark(self):
        if self.loading:return
        self.dirty=True
        try:self.set_preview(compact(self.result()));self.status.set("标注已修改，尚未保存")
        except ValueError as e:self.set_preview({"action":self.action.get()} if self.action.get() else "");self.status.set(f"尚未完成：{e}")
    def refresh_list(self):
        pass
    def canvas_changed(self):self.refresh_list();self.mark()
    def undo(self):
        if self.action.get()=="click" and self.canvas.boxes:self.canvas.boxes.pop();self.canvas.redraw();self.canvas_changed()
        elif self.action.get()=="swipe" and self.canvas.swipe:self.canvas.swipe=None;self.canvas.swipe_end=None;self.canvas.redraw();self.canvas_changed()
    def clear(self):self.canvas.boxes=[];self.canvas.swipe=None;self.canvas.swipe_end=None;self.canvas.redraw();self.canvas_changed()
    def save(self):
        if not self.dirty:self.status.set("当前条没有修改");return True
        try:
            text=compact(self.result());t=self.tasks[self.index];self.ws.cell(t.row,self.headers["结果输出"]).value=text;tmp=self.path.with_name(f".{self.path.stem}.annotation.tmp{self.path.suffix}")
            try:self.wb.save(tmp);os.replace(tmp,self.path)
            finally:
                if tmp.exists():tmp.unlink()
            self.tasks[self.index]=Task(t.row,t.instruction,t.image_id,t.xml,text);self.dirty=False;self.set_preview(text);self.status.set("已保存到结果输出列");return True
        except (OSError,ValueError) as e:messagebox.showerror("保存失败",str(e),parent=self.root);self.status.set(f"保存失败：{e}");return False
    def nav(self,d):
        target=self.index+d
        if not 0<=target<len(self.tasks):self.status.set("已经是第一条" if target<0 else "已经是最后一条");return
        if self.save():self.show(target)
    def list_selected(self,_event=None):
        if self.selecting or self.loading:return
        selected=self.task_list.selection()
        if not selected:return
        target=int(selected[0])
        if target==self.index:return
        if self.save():self.show(target)
        else:
            self.selecting=True;self.task_list.selection_set(str(self.index));self.selecting=False
    def close(self):
        if self.save():self.wb.close();self.root.destroy()

def main(argv=None):
    parser = argparse.ArgumentParser(description="Excel 用例参考答案 TTK 标注工具")
    parser.add_argument("workbook", nargs="?")
    args = parser.parse_args(argv)
    try:
        print("[ANNOTATION] 正在初始化 TTK...", flush=True)
        root = tk.Tk()
    except tk.TclError as error:
        print(f"无法初始化 TTK 界面：{error}", file=sys.stderr)
        return 1
    root.withdraw()
    print("[ANNOTATION] TTK 初始化完成。", flush=True)
    selected = args.workbook or filedialog.askopenfilename(
        title="选择评测任务表",
        filetypes=(("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")),
    )
    if not selected:
        print("[ANNOTATION] 未选择工作簿，程序退出。", flush=True)
        root.destroy()
        return 0
    selected = str(Path(selected).expanduser().resolve())
    print(f"[ANNOTATION] 正在打开工作簿：{selected}", flush=True)
    root.title("正在加载标注工具...")
    root.geometry("640x180")
    loading = ttk.Label(
        root,
        text=f"正在加载：\n{selected}",
        anchor="center",
        justify=tk.CENTER,
        padding=20,
    )
    loading.pack(fill=tk.BOTH, expand=True)
    root.deiconify()
    root.lift()
    root.update_idletasks()
    root.update()
    try:
        loading.destroy()
        App(root, Path(selected))
        print("[ANNOTATION] 主界面加载完成。", flush=True)
        root.lift()
        root.after(100, root.focus_force)
    except Exception as error:
        print(f"无法打开标注工具：{type(error).__name__}: {error}", file=sys.stderr)
        root.deiconify()
        root.lift()
        messagebox.showerror("无法打开标注工具", str(error), parent=root)
        root.destroy()
        return 1
    print("[ANNOTATION] 已进入界面事件循环。", flush=True)
    root.mainloop()
    return 0

if __name__=="__main__":sys.exit(main())
