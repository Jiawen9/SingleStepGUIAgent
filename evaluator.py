"""Offline Excel evaluation for the GUI action engines."""

from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor
import json
import os
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

from config import AgentConfig, load_env_file
from contracts import ActionSelection, DecisionOutcome, ExecutionInput, ScreenSnapshot
from engines.vla.prompts import load_app_prompt
from orchestrator import DEFAULT_ENGINE_ORDER, PROJECT_ROOT, Pipeline
from output.commands import CommandBuilder
from output.serialization import action_as_prompt_object
from storage.artifacts import prepare_run_artifacts, save_original_screenshot


TASK_SHEET = "测试用例集"
DETAIL_SHEET = "评测明细"
SUMMARY_SHEET = "总览"
REQUIRED_COLUMNS = ("任务指令", "图片ID", "UI-TREE", "二级能力", "三级能力", "结果输出")
ABILITY_CATEGORIES = ("文本-清晰", "文本-模糊", "图标-清晰", "图标-模糊", "拒答")
EVALUATOR_ENGINES = ("xml", "ocr", "vla")
VLA_MODES = ("vla-basic", "vla-combo")
ENGINE_STRATEGIES = ("serial", "parallel")


@dataclass(frozen=True)
class EvaluationRecord:
    source_row: int
    instruction: str
    image_id: str
    ui_tree: str
    secondary_ability: str
    tertiary_ability: str
    ability_category: str
    expected_json: str
    actual_json: str
    pixel_action_json: str
    selected_engine: str
    engine_trace: str
    correct: bool
    comparison: str
    decision_seconds: float
    engine_seconds: float
    error: str
    artifact_path: str
    engine_details_json: str
    engine_correct: dict[str, bool]


@dataclass(frozen=True)
class _EngineEvaluation:
    name: str
    action_payload: dict[str, Any] | None
    pixel_payload: dict[str, Any] | None
    selected: bool
    correct: bool
    comparison: str
    status: str
    decision_seconds: float
    engine_seconds: float
    error: str
    artifact: str


def _string(value: Any) -> str:
    return "" if value is None else str(value).strip()


def ability_category(secondary: str, tertiary: str) -> str:
    if secondary == "拒答":
        return "拒答"
    mapping = {
        ("文本定位", "意图清晰"): "文本-清晰",
        ("文本定位", "意图模糊"): "文本-模糊",
        ("图标定位", "意图清晰"): "图标-清晰",
        ("图标定位", "意图模糊"): "图标-模糊",
    }
    return mapping.get((secondary, tertiary), "未分类")


def _resolve_existing_path(value: str, *, workbook_dir: Path, project_root: Path) -> Path | None:
    candidate = Path(value).expanduser()
    candidates = [candidate] if candidate.is_absolute() else [workbook_dir / candidate, project_root / candidate]
    for item in candidates:
        if item.is_file():
            return item.resolve()
    return None


def resolve_image_path(image_id: str, *, workbook_dir: Path, project_root: Path) -> Path:
    value = Path(image_id).expanduser()
    candidate = value if value.is_absolute() else workbook_dir / "device_captures" / value
    candidate = candidate.resolve()
    if candidate.is_file():
        return candidate

    parent = candidate.parent
    similar = []
    if parent.is_dir():
        similar = [item.name for item in parent.glob(f"{candidate.stem}*")][:10]
    raise FileNotFoundError(
        f"Image was not found for 图片ID={image_id!r}. "
        f"Path: {candidate}; exists={candidate.exists()}; is_file={candidate.is_file()}; "
        f"parent_exists={parent.is_dir()}; similar_names={similar}"
    )


def load_ui_tree(value: str, *, workbook_dir: Path, project_root: Path) -> tuple[Path | None, ElementTree.Element | None]:
    if not value:
        return None, None
    if value.lstrip().startswith("<"):
        try:
            return None, ElementTree.fromstring(value)
        except ElementTree.ParseError as error:
            raise ValueError(f"Invalid inline UI-TREE XML: {error}") from error
    candidate = Path(value).expanduser()
    path = candidate if candidate.is_absolute() else workbook_dir / "device_captures" / candidate
    path = path.resolve()
    if not path.is_file():
        parent = path.parent
        similar = [item.name for item in parent.glob(f"{path.stem}*")][:10] if parent.is_dir() else []
        raise FileNotFoundError(
            f"UI-TREE file was not found: {value}. Path: {path}; "
            f"exists={path.exists()}; is_file={path.is_file()}; "
            f"parent_exists={parent.is_dir()}; similar_names={similar}"
        )
    try:
        return path, ElementTree.fromstring(path.read_text(encoding="utf-8"))
    except ElementTree.ParseError as error:
        raise ValueError(f"Invalid UI-TREE XML in {path}: {error}") from error


def detect_registered_app_package(xml_root: ElementTree.Element | None) -> str | None:
    if xml_root is None:
        return None
    for node in xml_root.iter():
        package = (node.get("package") or "").strip()
        if package and load_app_prompt(package) is not None:
            return package
    return None


def load_snapshot(path: Path) -> ScreenSnapshot:
    png = path.read_bytes()
    try:
        with Image.open(BytesIO(png)) as image:
            image.load()
            width, height = image.size
    except (OSError, ValueError) as error:
        raise ValueError(f"Unable to read image: {path}") from error
    if width <= 0 or height <= 0:
        raise ValueError(f"Image has invalid dimensions: {path}")
    return ScreenSnapshot(png, width, height, None)


def swipe_direction(arguments: dict[str, Any]) -> str:
    dx = float(arguments["x2"]) - float(arguments["x1"])
    dy = float(arguments["y2"]) - float(arguments["y1"])
    if abs(dx) >= abs(dy):
        return "right" if dx > 0 else "left"
    return "down" if dy > 0 else "up"


def compare_action(
    expected: dict[str, Any],
    actual: ActionSelection | None,
    command,
) -> tuple[bool, str]:
    expected_name = expected.get("action") or expected.get("action_id")
    if not isinstance(expected_name, str) or not expected_name:
        raise ValueError("结果输出 must contain action or action_id.")
    if actual is None:
        return False, "No engine selected an action."
    if actual.name != expected_name:
        return False, f"Action mismatch: expected {expected_name}, got {actual.name}."

    if expected_name == "click":
        boxes = expected.get("bbox")
        if not isinstance(boxes, list) or not boxes:
            raise ValueError("click result must contain a non-empty bbox list.")
        if command is None or command.target != "tap":
            return False, "Selected click did not produce a pixel tap command."
        x, y = command.arguments["x"], command.arguments["y"]
        for box in boxes:
            if (
                isinstance(box, list)
                and len(box) == 4
                and all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in box)
            ):
                x1, y1, x2, y2 = box
                if x1 <= x <= x2 and y1 <= y <= y2:
                    return True, f"Click pixel ({x}, {y}) is inside bbox {box}."
        return False, f"Click pixel ({x}, {y}) is outside all expected boxes."

    if expected_name == "swipe":
        expected_direction = expected.get("direction")
        if expected_direction not in {"up", "down", "left", "right"}:
            raise ValueError("swipe result must contain a valid direction.")
        actual_direction = swipe_direction(actual.arguments)
        return (
            actual_direction == expected_direction,
            f"Swipe direction: expected {expected_direction}, got {actual_direction}.",
        )

    if expected_name == "type":
        if set(expected) != {"action", "text"} or not isinstance(expected.get("text"), str):
            raise ValueError("type result must contain exactly action and text.")
        matches = actual.arguments.get("text") == expected["text"]
        return matches, "Type text matched." if matches else "Type text did not match."

    expected_arguments = {
        key: value for key, value in expected.items() if key not in {"action", "action_id"}
    }
    matches = actual.arguments == expected_arguments
    return (
        matches,
        "Action name and parameters matched."
        if matches
        else f"Parameter mismatch: expected {expected_arguments}, got {actual.arguments}.",
    )


def pixel_action_payload(outcome: DecisionOutcome) -> dict[str, Any] | None:
    command = outcome.command
    selected = outcome.selected_engine_result
    if command is None or selected is None or selected.action is None:
        return None
    if command.target == "tap":
        return {"action": "click", "coordinate": [command.arguments["x"], command.arguments["y"]]}
    if command.target == "swipe":
        return {
            "action": "swipe",
            "start_coordinate": [command.arguments["x1"], command.arguments["y1"]],
            "end_coordinate": [command.arguments["x2"], command.arguments["y2"]],
            "direction": swipe_direction(selected.action.arguments),
        }
    return action_as_prompt_object(selected.action)


def _uses_vla_combo(vla_mode: str) -> bool:
    return vla_mode == "vla-combo"


def create_engine_pipelines(
    config: AgentConfig, *, project_root: Path
) -> dict[str, Pipeline]:
    pipelines: dict[str, Pipeline] = {}
    for engine_name in EVALUATOR_ENGINES:
        pipeline = Pipeline(config=config, project_root=project_root, engine_order=DEFAULT_ENGINE_ORDER)
        # Offline evaluation compares the selected action and never executes
        # device tools. App actions without an implementation must therefore
        # remain valid decision results instead of failing command conversion.
        pipeline.command_builder = CommandBuilder(allow_unmapped=True)
        pipeline.engines = tuple(
            engine for engine in pipeline.engines if engine.name == engine_name
        )
        pipelines[engine_name] = pipeline
    return pipelines


def _write_full_decision(path: Path, outcome: DecisionOutcome) -> None:
    payload = {
        "engine_results": [result.as_dict() for result in outcome.engine_results],
        "selected_engine_result": (
            outcome.selected_engine_result.as_dict()
            if outcome.selected_engine_result is not None
            else None
        ),
        "command": outcome.command.as_dict() if outcome.command is not None else None,
        "timings_seconds": outcome.timings_seconds,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _evaluate_engine(
    *,
    engine_name: str,
    engine_pipeline: Pipeline,
    row_number: int,
    instruction: str,
    snapshot: ScreenSnapshot,
    app_package: str | None,
    xml_path: Path | None,
    xml_root: ElementTree.Element | None,
    image_path: Path,
    project_root: Path,
    run_id: str,
    vla_mode: str,
    engine_strategy: str,
    expected: dict[str, Any],
) -> _EngineEvaluation:
    print(f"[EVAL] ENGINE START row={row_number} engine={engine_name}", flush=True)
    case_id = f"EVAL-{run_id}-{row_number:04d}-{engine_name}"
    paths = prepare_run_artifacts(project_root=project_root, case_id=case_id)
    save_original_screenshot(paths=paths, snapshot=snapshot)
    execution_input = ExecutionInput(
        case_id,
        instruction,
        snapshot,
        app_package if engine_name == "vla" else None,
        paths.directory,
        xml_path,
        xml_root,
        {
            "source": str(image_path),
            "evaluation_mode": "xml-ocr-vla",
            "evaluation_vla_mode": vla_mode,
            "evaluation_engine": engine_name,
            "evaluation_engine_strategy": engine_strategy,
        },
    )
    try:
        outcome = engine_pipeline.decide(execution_input, paths=paths)
        decision_seconds = outcome.timings_seconds.get("decision", 0.0)
        engine_seconds = outcome.timings_seconds.get("engines", 0.0)
        decision_path = paths.directory / "decision.json"
        _write_full_decision(decision_path, outcome)
        artifact = str(decision_path.relative_to(project_root))
        selected = outcome.selected_engine_result
        action_payload = (
            action_as_prompt_object(selected.action)
            if selected is not None and selected.action is not None
            else None
        )
        pixel_payload = pixel_action_payload(outcome)
        engine_ok, comparison = compare_action(
            expected,
            selected.action if selected is not None else None,
            outcome.command,
        )
        last_result = outcome.engine_results[-1] if outcome.engine_results else None
        status = last_result.status if last_result is not None else "not_run"
        engine_error = ""
        if last_result is not None and last_result.status == "error":
            engine_error = str(last_result.diagnostics.get("error", "Engine error."))
        print(
            f"[EVAL] ENGINE {'PASS' if engine_ok else 'FAIL'} "
            f"row={row_number} engine={engine_name} elapsed={decision_seconds:.3f}s",
            flush=True,
        )
        return _EngineEvaluation(
            engine_name,
            action_payload,
            pixel_payload,
            action_payload is not None,
            engine_ok,
            comparison,
            status,
            decision_seconds,
            engine_seconds,
            engine_error,
            artifact,
        )
    except Exception as exception:
        error_text = f"{type(exception).__name__}: {exception}"
        print(
            f"[EVAL] ENGINE ERROR row={row_number} engine={engine_name} error={error_text}",
            flush=True,
        )
        return _EngineEvaluation(
            engine_name,
            None,
            None,
            False,
            False,
            "Evaluation error.",
            "error",
            0.0,
            0.0,
            error_text,
            "",
        )


def _evaluate_row(
    row_number: int,
    row_values: dict[str, str],
    *,
    pipelines: dict[str, Pipeline],
    project_root: Path,
    workbook_dir: Path,
    run_id: str,
    vla_mode: str,
    engine_strategy: str,
) -> EvaluationRecord:
    instruction = row_values["任务指令"]
    image_id = row_values["图片ID"]
    ui_tree_value = row_values["UI-TREE"]
    secondary = row_values["二级能力"]
    tertiary = row_values["三级能力"]
    category = ability_category(secondary, tertiary)
    expected_json = row_values["结果输出"]
    actual_json = pixel_json = selected_engine = trace = comparison = error = artifact = ""
    engine_details_json = "{}"
    engine_correct: dict[str, bool] = {}
    correct = False
    decision_seconds = engine_seconds = 0.0
    try:
        if not instruction or not image_id or not expected_json:
            raise ValueError("任务指令、图片ID和结果输出不能为空。")
        expected = json.loads(expected_json)
        if not isinstance(expected, dict):
            raise ValueError("结果输出 must be a JSON object.")
        image_path = resolve_image_path(image_id, workbook_dir=workbook_dir, project_root=project_root)
        snapshot = load_snapshot(image_path)
        xml_path, xml_root = load_ui_tree(
            ui_tree_value, workbook_dir=workbook_dir, project_root=project_root
        )
        app_package = (
            detect_registered_app_package(xml_root)
            if _uses_vla_combo(vla_mode)
            else None
        )
        engine_items = list(pipelines.items())

        def run_engine(item: tuple[str, Pipeline]) -> _EngineEvaluation:
            engine_name, engine_pipeline = item
            return _evaluate_engine(
                engine_name=engine_name,
                engine_pipeline=engine_pipeline,
                row_number=row_number,
                instruction=instruction,
                snapshot=snapshot,
                app_package=app_package,
                xml_path=xml_path,
                xml_root=xml_root,
                image_path=image_path,
                project_root=project_root,
                run_id=run_id,
                vla_mode=vla_mode,
                engine_strategy=engine_strategy,
                expected=expected,
            )

        parallel_union = engine_strategy == "parallel" and len(engine_items) > 1
        if parallel_union:
            parallel_started = time.perf_counter()
            with ThreadPoolExecutor(max_workers=len(engine_items)) as executor:
                evaluations = list(executor.map(run_engine, engine_items))
            parallel_wall_seconds = time.perf_counter() - parallel_started
        else:
            evaluations = []
            for item in engine_items:
                evaluation = run_engine(item)
                evaluations.append(evaluation)
                # Serial mode is a fallback chain. A selected action is a hit
                # even when it is later judged wrong; only no_match/error falls
                # through to the next engine.
                if evaluation.selected:
                    break

        details: dict[str, dict[str, Any]] = {}
        actual_by_engine: dict[str, Any] = {}
        pixel_by_engine: dict[str, Any] = {}
        selected_engines: list[str] = []
        traces: list[str] = []
        comparisons: list[str] = []
        artifacts: dict[str, str] = {}
        for evaluation in evaluations:
            engine_correct[evaluation.name] = evaluation.correct
            actual_by_engine[evaluation.name] = evaluation.action_payload
            pixel_by_engine[evaluation.name] = evaluation.pixel_payload
            if evaluation.selected:
                selected_engines.append(evaluation.name)
            if evaluation.artifact:
                artifacts[evaluation.name] = evaluation.artifact
            traces.append(f"{evaluation.name}:{evaluation.status}")
            if evaluation.error:
                comparisons.append(f"{evaluation.name}=ERROR: {evaluation.error}")
            else:
                comparisons.append(
                    f"{evaluation.name}={'PASS' if evaluation.correct else 'FAIL'}: "
                    f"{evaluation.comparison}"
                )
            details[evaluation.name] = {
                "action": evaluation.action_payload,
                "pixel_action": evaluation.pixel_payload,
                "correct": evaluation.correct,
                "comparison": evaluation.comparison,
                "decision_seconds": evaluation.decision_seconds,
                "engine_seconds": evaluation.engine_seconds,
                "error": evaluation.error,
                "artifact": evaluation.artifact,
            }

        if parallel_union:
            correct = any(evaluation.correct for evaluation in evaluations)
            actual_json = json.dumps(
                actual_by_engine, ensure_ascii=False, separators=(",", ":")
            )
            pixel_json = json.dumps(
                pixel_by_engine, ensure_ascii=False, separators=(",", ":")
            )
            selected_engine = ",".join(selected_engines)
            comparison = " | ".join(comparisons)
            artifact = json.dumps(artifacts, ensure_ascii=False, separators=(",", ":"))
            decision_seconds = parallel_wall_seconds
        else:
            selected_evaluation = next(
                (evaluation for evaluation in evaluations if evaluation.selected), None
            )
            correct = selected_evaluation.correct if selected_evaluation is not None else False
            actual_json = (
                json.dumps(
                    selected_evaluation.action_payload,
                    ensure_ascii=False,
                    separators=(",", ":"),
                )
                if selected_evaluation is not None
                else ""
            )
            pixel_json = (
                json.dumps(
                    selected_evaluation.pixel_payload,
                    ensure_ascii=False,
                    separators=(",", ":"),
                )
                if selected_evaluation is not None
                and selected_evaluation.pixel_payload is not None
                else ""
            )
            selected_engine = selected_evaluation.name if selected_evaluation is not None else ""
            comparison = (
                selected_evaluation.comparison
                if selected_evaluation is not None
                else " | ".join(comparisons)
            )
            artifact = selected_evaluation.artifact if selected_evaluation is not None else ""
            decision_seconds = sum(
                evaluation.decision_seconds for evaluation in evaluations
            )
        engine_seconds = sum(evaluation.engine_seconds for evaluation in evaluations)
        trace = " | ".join(traces)
        engine_details_json = json.dumps(details, ensure_ascii=False, separators=(",", ":"))
        engine_errors = {
            evaluation.name: evaluation.error
            for evaluation in evaluations
            if evaluation.error
        }
        if selected_engine:
            error = ""
        elif len(engine_errors) == 1:
            error = next(iter(engine_errors.values()))
        elif engine_errors:
            error = " | ".join(f"{name}: {message}" for name, message in engine_errors.items())
    except Exception as exception:
        error = f"{type(exception).__name__}: {exception}"
        comparison = "Evaluation error."
    return EvaluationRecord(
        row_number, instruction, image_id, ui_tree_value, secondary, tertiary, category, expected_json,
        actual_json, pixel_json, selected_engine, trace, correct, comparison,
        decision_seconds, engine_seconds, error, artifact, engine_details_json, engine_correct,
    )


def evaluate_workbook(
    input_path: Path,
    *,
    output_path: Path,
    pipeline: dict[str, Pipeline],
    project_root: Path,
    vla_mode: str = "vla-combo",
    workers: int = 1,
    engine_strategy: str = "serial",
) -> list[EvaluationRecord]:
    if workers < 1:
        raise ValueError("workers must be at least 1.")
    if engine_strategy not in ENGINE_STRATEGIES:
        raise ValueError(
            f"engine_strategy must be one of: {', '.join(ENGINE_STRATEGIES)}."
        )
    if vla_mode not in VLA_MODES:
        raise ValueError(f"vla_mode must be one of: {', '.join(VLA_MODES)}.")
    workbook = load_workbook(input_path)
    if TASK_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook must contain a {TASK_SHEET!r} sheet.")
    task_sheet = workbook[TASK_SHEET]
    headers = {_string(cell.value): index for index, cell in enumerate(task_sheet[1], start=1)}
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ValueError("Missing task columns: " + ", ".join(missing))

    if not isinstance(pipeline, dict):
        raise ValueError(
            "Evaluator requires independent xml, ocr, and vla pipelines."
        )
    missing_pipelines = [name for name in EVALUATOR_ENGINES if name not in pipeline]
    if missing_pipelines:
        raise ValueError("Missing evaluator pipelines: " + ", ".join(missing_pipelines))
    pipelines = {name: pipeline[name] for name in EVALUATOR_ENGINES}

    run_id = datetime.now().strftime("%Y%m%d-%H%M%S")
    workbook_dir = input_path.resolve().parent
    tasks: list[tuple[int, dict[str, str]]] = []
    for row_number in range(2, task_sheet.max_row + 1):
        row_values = {
            name: _string(task_sheet.cell(row=row_number, column=headers[name]).value)
            for name in REQUIRED_COLUMNS
        }
        if not any(row_values.values()):
            continue
        tasks.append((row_number, row_values))

    total = len(tasks)
    print(
        f"[EVAL] 已加载 {total} 条用例，engines=xml,ocr,vla，vla_mode={vla_mode}，"
        f"engine_strategy={engine_strategy}，workers={workers}",
        flush=True,
    )
    if not tasks:
        print("[EVAL] 没有可执行的非空用例。", flush=True)

    progress_lock = threading.Lock()
    completed = 0

    def evaluate_task(task: tuple[int, dict[str, str]]) -> EvaluationRecord:
        nonlocal completed
        row_number, row_values = task
        instruction = row_values["任务指令"].replace("\r", " ").replace("\n", " ")
        with progress_lock:
            print(
                f"[EVAL] START row={row_number} image={row_values['图片ID']} "
                f"instruction={instruction}",
                flush=True,
            )
        started = time.perf_counter()
        record = _evaluate_row(
            row_number,
            row_values,
            pipelines=pipelines,
            project_root=project_root,
            workbook_dir=workbook_dir,
            run_id=run_id,
            vla_mode=vla_mode,
            engine_strategy=engine_strategy,
        )
        elapsed = time.perf_counter() - started
        status = "ERROR" if record.error else ("PASS" if record.correct else "FAIL")
        with progress_lock:
            completed += 1
            detail = f" error={record.error}" if record.error else ""
            print(
                f"[EVAL] {status} row={row_number} progress={completed}/{total} "
                f"engine={record.selected_engine or '-'} elapsed={elapsed:.3f}s{detail}",
                flush=True,
            )
        return record

    if workers == 1:
        records = [evaluate_task(task) for task in tasks]
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            records = list(executor.map(evaluate_task, tasks))
    records.sort(key=lambda record: record.source_row)

    print(f"[EVAL] 正在写入结果工作簿：{output_path}", flush=True)
    write_result_workbook(
        workbook,
        records,
        output_path=output_path,
        vla_mode=vla_mode,
        engine_strategy=engine_strategy,
    )
    print(f"[EVAL] 评测完成，结果已写入：{output_path}", flush=True)
    return records


def _style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in row:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_result_workbook(
    workbook,
    records: list[EvaluationRecord],
    *,
    output_path: Path,
    vla_mode: str,
    engine_strategy: str,
) -> None:
    for name in (DETAIL_SHEET, SUMMARY_SHEET):
        if name in workbook.sheetnames:
            del workbook[name]
    detail = workbook.create_sheet(DETAIL_SHEET)
    headers = [
        "源行号", "任务指令", "图片ID", "UI-TREE", "二级能力", "三级能力", "能力分类",
        "期望结果", "系统标准动作", "系统像素动作", "执行主体", "Engine状态链", "是否正确", "判分说明", "决策时延(秒)",
        "Engine时延(秒)", "错误", "完整决策留档", "各Engine评测结果",
        "XML正确", "OCR正确", "VLA正确",
    ]
    detail.append(headers)
    for record in records:
        detail.append([
            record.source_row, record.instruction, record.image_id, record.ui_tree,
            record.secondary_ability, record.tertiary_ability, record.ability_category,
            record.expected_json, record.actual_json, record.pixel_action_json, record.selected_engine,
            record.engine_trace, record.correct, record.comparison, record.decision_seconds,
            record.engine_seconds, record.error, record.artifact_path, record.engine_details_json,
            record.engine_correct.get("xml", ""), record.engine_correct.get("ocr", ""),
            record.engine_correct.get("vla", ""),
        ])
    _style_header(detail[1])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    widths = [10, 24, 16, 34, 16, 16, 16, 44, 46, 46, 18, 34, 12, 60, 16, 16, 42, 48, 80, 12, 12, 12]
    for index, width in enumerate(widths, start=1):
        detail.column_dimensions[get_column_letter(index)].width = width
    for row in detail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in (15, 16):
        for cell in detail.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = "0.000"
    if records:
        result_range = f"M2:M{len(records) + 1}"
        detail.conditional_formatting.add(
            result_range,
            CellIsRule(operator="equal", formula=["TRUE"], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        detail.conditional_formatting.add(
            result_range,
            CellIsRule(operator="equal", formula=["FALSE"], fill=PatternFill("solid", fgColor="FFC7CE")),
        )

    summary = workbook.create_sheet(SUMMARY_SHEET, 0)
    summary.append(["GUI Agent 批量评测总览", ""])
    summary.merge_cells("A1:B1")
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.append(["评测引擎", "XML+OCR+VLA"])
    summary["D2"] = "引擎执行策略"
    summary["E2"] = engine_strategy
    summary.append(["VLA动作模式", vla_mode])
    summary.append(["生成时间", datetime.now()])
    summary.append(["任务总数", f"=COUNTA('{DETAIL_SHEET}'!A2:A{max(2, len(records) + 1)})"])
    summary.append(["正确数", f"=COUNTIF('{DETAIL_SHEET}'!M2:M{max(2, len(records) + 1)},TRUE)"])
    summary.append(["失败数", "=B5-B6"])
    summary.append(["错误数", f'=COUNTIF(\'{DETAIL_SHEET}\'!Q2:Q{max(2, len(records) + 1)},"?*")'])
    summary.append(["总成功率", "=IFERROR(B6/B5,0)"])
    summary.append(["平均决策时延(秒)", f"=IFERROR(AVERAGE('{DETAIL_SHEET}'!O2:O{max(2, len(records) + 1)}),0)"])
    summary["B4"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary["B9"].number_format = "0.0%"
    summary["B10"].number_format = "0.000"

    summary.append([])
    summary.append(["执行主体", "执行数量", "正确数量", "成功率"])
    _style_header(summary[12])
    engine_columns = {"xml": "T", "ocr": "U", "vla": "V"}
    for source, column in engine_columns.items():
        row = summary.max_row + 1
        result_range = f"'{DETAIL_SHEET}'!{column}2:{column}{max(2, len(records) + 1)}"
        summary.append([
            source,
            f'=COUNTIF({result_range},TRUE)+COUNTIF({result_range},FALSE)',
            f'=COUNTIF({result_range},TRUE)',
            f"=IFERROR(C{row}/B{row},0)",
        ])
        summary.cell(row, 4).number_format = "0.0%"

    summary.append([])
    grouped_parallel_summary = engine_strategy == "parallel"
    ability_headers = [
        "能力分类", "任务数", "综合正确数", "综合成功率",
        "XML正确数", "XML成功率", "OCR正确数", "OCR成功率", "VLA正确数", "VLA成功率",
    ]
    if grouped_parallel_summary:
        ability_headers.extend(["XML+OCR并集正确数", "XML+OCR并集成功率"])
    summary.append(ability_headers)
    _style_header(summary[summary.max_row])
    detail_last_row = max(2, len(records) + 1)
    for category in (*ABILITY_CATEGORIES, "总体"):
        row = summary.max_row + 1
        if category == "总体":
            category_condition = None
            task_count = f"=COUNTA('{DETAIL_SHEET}'!A2:A{detail_last_row})"
            combined_correct = f"=COUNTIF('{DETAIL_SHEET}'!M2:M{detail_last_row},TRUE)"
        else:
            category_condition = f"'{DETAIL_SHEET}'!G2:G{detail_last_row},A{row}"
            task_count = f"=COUNTIF('{DETAIL_SHEET}'!G2:G{detail_last_row},A{row})"
            combined_correct = (
                f"=COUNTIFS('{DETAIL_SHEET}'!G2:G{detail_last_row},A{row},"
                f"'{DETAIL_SHEET}'!M2:M{detail_last_row},TRUE)"
            )
        values: list[Any] = [category, task_count, combined_correct, f"=IFERROR(C{row}/B{row},0)"]
        for column in ("T", "U", "V"):
            if category_condition is None:
                correct_formula = f"=COUNTIF('{DETAIL_SHEET}'!{column}2:{column}{detail_last_row},TRUE)"
                executed_formula = (
                    f"COUNTIF('{DETAIL_SHEET}'!{column}2:{column}{detail_last_row},TRUE)+"
                    f"COUNTIF('{DETAIL_SHEET}'!{column}2:{column}{detail_last_row},FALSE)"
                )
            else:
                correct_formula = (
                    f"=COUNTIFS('{DETAIL_SHEET}'!G2:G{detail_last_row},A{row},"
                    f"'{DETAIL_SHEET}'!{column}2:{column}{detail_last_row},TRUE)"
                )
                executed_formula = (
                    f"COUNTIFS('{DETAIL_SHEET}'!G2:G{detail_last_row},A{row},"
                    f"'{DETAIL_SHEET}'!{column}2:{column}{detail_last_row},TRUE)+"
                    f"COUNTIFS('{DETAIL_SHEET}'!G2:G{detail_last_row},A{row},"
                    f"'{DETAIL_SHEET}'!{column}2:{column}{detail_last_row},FALSE)"
                )
            correct_column = get_column_letter(len(values) + 1)
            rate_denominator = f"B{row}" if grouped_parallel_summary else f"({executed_formula})"
            values.extend([
                correct_formula,
                f"=IFERROR({correct_column}{row}/{rate_denominator},0)",
            ])
        if grouped_parallel_summary:
            if category_condition is None:
                xml_ocr_union_correct = (
                    f"=COUNTIF('{DETAIL_SHEET}'!T2:T{detail_last_row},TRUE)+"
                    f"COUNTIF('{DETAIL_SHEET}'!U2:U{detail_last_row},TRUE)-"
                    f"COUNTIFS('{DETAIL_SHEET}'!T2:T{detail_last_row},TRUE,"
                    f"'{DETAIL_SHEET}'!U2:U{detail_last_row},TRUE)"
                )
            else:
                xml_ocr_union_correct = (
                    f"=COUNTIFS('{DETAIL_SHEET}'!G2:G{detail_last_row},A{row},"
                    f"'{DETAIL_SHEET}'!T2:T{detail_last_row},TRUE)+"
                    f"COUNTIFS('{DETAIL_SHEET}'!G2:G{detail_last_row},A{row},"
                    f"'{DETAIL_SHEET}'!U2:U{detail_last_row},TRUE)-"
                    f"COUNTIFS('{DETAIL_SHEET}'!G2:G{detail_last_row},A{row},"
                    f"'{DETAIL_SHEET}'!T2:T{detail_last_row},TRUE,"
                    f"'{DETAIL_SHEET}'!U2:U{detail_last_row},TRUE)"
                )
            union_column = get_column_letter(len(values) + 1)
            values.extend([
                xml_ocr_union_correct,
                f"=IFERROR({union_column}{row}/B{row},0)",
            ])
        summary.append(values)
        percentage_columns = (4, 6, 8, 10, 12) if grouped_parallel_summary else (4, 6, 8, 10)
        for column in percentage_columns:
            summary.cell(row, column).number_format = "0.0%"
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 22
    summary.column_dimensions["C"].width = 16
    summary.column_dimensions["D"].width = 16
    summary.column_dimensions["E"].width = 22
    for column in ("F", "G", "H", "I", "J"):
        summary.column_dimensions[column].width = 16
    if grouped_parallel_summary:
        summary.column_dimensions["K"].width = 22
        summary.column_dimensions["L"].width = 22
    summary.freeze_panes = "A2"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_config() -> AgentConfig:
    api_base = os.environ.get("MODEL_URL") or os.environ.get("YUNAI_API_BASE")
    model = os.environ.get("MODEL_NAME") or os.environ.get("YUNAI_MODEL")
    if not api_base:
        raise ValueError("MODEL_URL is required for VLA evaluation.")
    if not model:
        raise ValueError("MODEL_NAME is required for VLA evaluation.")
    return AgentConfig.from_values(
        api_key=os.environ.get("MODEL_API_KEY", os.environ.get("YUNAI_API_KEY", "")),
        api_base=api_base or "http://unused.invalid/v1",
        model=model or "unused",
        adb_path=os.environ.get("ADB_PATH") or "adb",
        device_id=None,
    )


def positive_worker_count(value: str) -> int:
    try:
        workers = int(value)
    except ValueError as error:
        raise argparse.ArgumentTypeError("workers must be an integer.") from error
    if workers < 1:
        raise argparse.ArgumentTypeError("workers must be at least 1.")
    return workers


def parse_arguments(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate GUI engines from an Excel task table.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--vla-mode",
        choices=VLA_MODES,
        default="vla-combo",
        help="VLA action space; XML, OCR, and VLA are always evaluated.",
    )
    parser.add_argument("--output", type=Path)
    parser.add_argument("--workers", type=positive_worker_count, default=1)
    parser.add_argument(
        "--engine-strategy",
        choices=ENGINE_STRATEGIES,
        default="serial",
        help=(
            "How engines run inside one case: serial uses fallback order; "
            "parallel runs every engine and unions correct results."
        ),
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    try:
        load_env_file(PROJECT_ROOT / ".env")
        args = parse_arguments(argv)
        input_path = args.workbook.expanduser().resolve()
        if not input_path.is_file():
            raise FileNotFoundError(f"Workbook was not found: {input_path}")
        output_path = args.output
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            output_path = input_path.with_name(f"{input_path.stem}_result_{timestamp}.xlsx")
        else:
            output_path = output_path.expanduser().resolve()
        config = build_config()
        pipeline = create_engine_pipelines(config, project_root=PROJECT_ROOT)
        started = time.perf_counter()
        records = evaluate_workbook(
            input_path,
            output_path=output_path,
            pipeline=pipeline,
            project_root=PROJECT_ROOT,
            vla_mode=args.vla_mode,
            workers=args.workers,
            engine_strategy=args.engine_strategy,
        )
        correct = sum(record.correct for record in records)
        rate = correct / len(records) if records else 0.0
        average = sum(record.decision_seconds for record in records) / len(records) if records else 0.0
        print(f"任务数：{len(records)}")
        print(f"正确数：{correct}")
        print(f"总成功率：{rate:.1%}")
        print(f"平均决策时延：{average:.3f} 秒")
        print(f"用例并发数：{args.workers}")
        print(f"引擎执行策略：{args.engine_strategy}")
        print(f"VLA动作模式：{args.vla_mode}")
        print(f"总耗时：{time.perf_counter() - started:.3f} 秒")
        print(f"结果工作簿：{output_path}")
        return 0
    except (OSError, RuntimeError, ValueError) as error:
        print(f"错误：{error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
