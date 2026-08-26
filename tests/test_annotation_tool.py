from __future__ import annotations
import tempfile, unittest
from pathlib import Path
from PIL import Image
from annotation_tool_ttk import App, actions_for_ui_tree, make_result, parse_result, resolve_image_path, select_task_sheet, swipe_details
from openpyxl import Workbook

class AnnotationLogicTests(unittest.TestCase):
    def test_action_summary_for_right_hand_list(self):
        self.assertEqual(App.action_summary('{"action":"player_pause"}'),'{"action":"player_pause"}')
        self.assertEqual(App.action_summary('{"action":"click","bbox":[[1,2,3,4]]}'),'{"action":"click","bbox":[[1,2,3,4]]}')
    def test_sheet_is_auto_detected_by_headers(self):
        workbook=Workbook();sheet=workbook.active;sheet.title="Sheet1";sheet.append(["二级能力","三级能力","任务指令","图片ID","UI-TREE","结果输出"])
        self.assertEqual(select_task_sheet(workbook).title,"Sheet1")
    def test_standard_sheet_name_is_preferred(self):
        workbook=Workbook();fallback=workbook.active;fallback.append(["二级能力","三级能力","任务指令","图片ID","UI-TREE","结果输出"])
        standard=workbook.create_sheet("测试用例集");standard.append(["任务指令","图片ID","UI-TREE","二级能力","三级能力","结果输出"])
        self.assertEqual(select_task_sheet(workbook).title,"测试用例集")
    def test_legacy_sheet_and_columns_are_supported(self):
        workbook=Workbook();sheet=workbook.active;sheet.title="测试用例表"
        sheet.append(["二级场景","任务指令","图片ID","UI-TREE","结果输出"])
        self.assertEqual(select_task_sheet(workbook).title,"测试用例表")
    def test_image_extension_resolution(self):
        with tempfile.TemporaryDirectory() as directory:
            root=Path(directory);Image.new("RGB",(10,10)).save(root/"CASE.png")
            self.assertEqual(resolve_image_path("CASE",root),(root/"CASE.png").resolve())
    def test_registered_and_unknown_packages(self):
        specs,label=actions_for_ui_tree('<node package="com.qiyi.video.pad"/>',Path("."))
        self.assertIn("player_pause",{x.name for x in specs});self.assertIn("com.qiyi.video.pad",label)
        specs,_=actions_for_ui_tree('<node package="unknown.app"/>',Path("."))
        self.assertEqual({x.name for x in specs},{"click","swipe","type","reject"})
        self.assertNotIn("player_pause",{x.name for x in specs})
    def test_relative_xml_is_resolved_from_capture_directory(self):
        with tempfile.TemporaryDirectory() as directory:
            root=Path(directory);(root/"CASE.xml").write_text('<node package="com.qiyi.video.pad"/>',encoding="utf-8")
            specs,_=actions_for_ui_tree("CASE.xml",root)
            self.assertIn("player_pause",{x.name for x in specs})
    def test_swipe_direction_and_distance(self):
        self.assertEqual(swipe_details((500,500),(500,350),1000,1000),("up","short"))
        self.assertEqual(swipe_details((100,100),(450,100),1000,1000),("right","medium"))
        self.assertEqual(swipe_details((900,100),(100,100),1000,1000),("left","long"))
    def test_result_protocols(self):
        self.assertEqual(make_result("click",[[1,2,30,40]]),{"action":"click","bbox":[[1,2,30,40]]})
        swipe={"start_coordinate":[10,20],"direction":"down","distance":"long"}
        self.assertEqual(make_result("swipe",swipe=swipe),{"action":"swipe",**swipe})
        self.assertEqual(make_result("reject", parameters={}), {"action":"reject"})
        with self.assertRaises(ValueError):
            parse_result('{"action_id":"reject"}')
        with self.assertRaises(ValueError):
            parse_result('{"action":"reject","reason":"hidden"}')
        self.assertEqual(parse_result('{"action":"type","text":"测试"}')["text"],"测试")
        with self.assertRaises(ValueError):
            parse_result('{"action":"swipe","direction":"up"}')

if __name__ == "__main__": unittest.main()
