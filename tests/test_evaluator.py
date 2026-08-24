from __future__ import annotations

import json
import tempfile
import threading
import time
import unittest
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from PIL import Image

from contracts import (
    ActionSelection,
    DecisionOutcome,
    EngineResult,
    ExecutionCommand,
)
from evaluator import (
    DETAIL_SHEET,
    SUMMARY_SHEET,
    TASK_SHEET,
    _uses_vla_combo,
    ability_category,
    compare_action,
    detect_registered_app_package,
    evaluate_workbook,
    load_ui_tree,
    parse_arguments,
    resolve_image_path,
    swipe_direction,
)


class StaticPipeline:
    def __init__(self, outcome):
        self.outcome = outcome
        self.inputs = []

    def decide(self, execution_input, *, paths=None):
        self.inputs.append(execution_input)
        return self.outcome


class ConcurrentPipeline:
    def __init__(self, outcome):
        self.outcome = outcome
        self.active = 0
        self.peak_active = 0
        self.completion_order = []
        self.lock = threading.Lock()

    def decide(self, execution_input, *, paths=None):
        with self.lock:
            self.active += 1
            self.peak_active = max(self.peak_active, self.active)
        try:
            if execution_input.instruction == "失败":
                raise RuntimeError("planned failure")
            delay = {"慢": 0.12, "中": 0.06, "快": 0.01}[execution_input.instruction]
            time.sleep(delay)
            with self.lock:
                self.completion_order.append(execution_input.instruction)
            return self.outcome
        finally:
            with self.lock:
                self.active -= 1


class ParseFailurePipeline:
    def decide(self, execution_input, *, paths=None):
        if execution_input.instruction == "解析失败":
            raise KeyError("model action is missing")
        return click_outcome(pixel_x=100, pixel_y=50)


class EngineOverlapTracker:
    def __init__(self):
        self.active = 0
        self.peak_active = 0
        self.lock = threading.Lock()


class DelayedEnginePipeline(StaticPipeline):
    def __init__(self, outcome, tracker, delay=0.05):
        super().__init__(outcome)
        self.tracker = tracker
        self.delay = delay

    def decide(self, execution_input, *, paths=None):
        self.inputs.append(execution_input)
        with self.tracker.lock:
            self.tracker.active += 1
            self.tracker.peak_active = max(
                self.tracker.peak_active, self.tracker.active
            )
        try:
            time.sleep(self.delay)
            return self.outcome
        finally:
            with self.tracker.lock:
                self.tracker.active -= 1


def click_outcome(x=500, y=500, pixel_x=100, pixel_y=50, source="vla"):
    action = ActionSelection("click", {"x": x, "y": y})
    selected = EngineResult("selected", source, action=action, timings_seconds={"engine": 0.2})
    command = ExecutionCommand(
        "adb",
        "tap",
        {"x": pixel_x, "y": pixel_y, "vla_x": x, "vla_y": y},
        action,
    )
    return DecisionOutcome((selected,), selected, command, {"engines": 0.2, "decision": 0.25})


def no_match_outcome(source):
    result = EngineResult("no_match", source, timings_seconds={"engine": 0.1})
    return DecisionOutcome((result,), None, None, {"engines": 0.1, "decision": 0.12})


def three_engine_pipelines(*, xml=None, ocr=None, vla=None):
    return {
        "xml": xml or StaticPipeline(no_match_outcome("xml")),
        "ocr": ocr or StaticPipeline(no_match_outcome("ocr")),
        "vla": vla or StaticPipeline(no_match_outcome("vla")),
    }


class EvaluationComparisonTests(unittest.TestCase):
    def test_click_accepts_any_box_and_inclusive_boundary(self):
        expected = {"action": "click", "bbox": [[0, 0, 10, 10], [100, 50, 120, 80]]}
        outcome = click_outcome(pixel_x=100, pixel_y=50)
        correct, _ = compare_action(
            expected, outcome.selected_engine_result.action, outcome.command
        )
        self.assertTrue(correct)

    def test_click_outside_boxes_fails(self):
        outcome = click_outcome(pixel_x=99, pixel_y=50)
        correct, reason = compare_action(
            {"action": "click", "bbox": [[100, 50, 120, 80]]},
            outcome.selected_engine_result.action,
            outcome.command,
        )
        self.assertFalse(correct)
        self.assertIn("outside", reason)

    def test_swipe_compares_only_direction(self):
        action = ActionSelection("swipe", {"x1": 500, "y1": 800, "x2": 510, "y2": 200})
        self.assertEqual(swipe_direction(action.arguments), "up")
        correct, _ = compare_action({"action": "swipe", "direction": "up"}, action, None)
        self.assertTrue(correct)

    def test_swipe_reference_accepts_pixel_start_and_distance(self):
        action = ActionSelection("swipe", {"x1": 500, "y1": 800, "x2": 510, "y2": 200})
        correct, _ = compare_action(
            {"action": "swipe", "start_coordinate": [1200, 900], "direction": "up", "distance": "long"},
            action,
            None,
        )
        self.assertTrue(correct)

    def test_type_is_exact(self):
        self.assertTrue(
            compare_action(
                {"action": "type", "text": "爱奇艺"},
                ActionSelection("type", {"text": "爱奇艺"}),
                None,
            )[0]
        )
        self.assertFalse(
            compare_action(
                {"action": "type", "text": "爱奇艺"},
                ActionSelection("type", {"text": "爱奇艺 "}),
                None,
            )[0]
        )

    def test_atomic_action_compares_all_parameters(self):
        expected = {"action": "player_set_quality", "quality": "720p"}
        self.assertTrue(
            compare_action(
                expected,
                ActionSelection("player_set_quality", {"quality": "720p"}),
                None,
            )[0]
        )
        self.assertFalse(
            compare_action(
                expected,
                ActionSelection("player_set_quality", {"quality": "1080p"}),
                None,
            )[0]
        )


class EvaluationWorkbookTests(unittest.TestCase):
    def test_detects_registered_app_after_unregistered_shell_package(self):
        root = ElementTree.fromstring(
            '<hierarchy><node package="com.smarthome.centerapp"/>'
            '<node package="com.netease.cloudmusic.iot"/></hierarchy>'
        )
        self.assertEqual(
            detect_registered_app_package(root),
            "com.netease.cloudmusic.iot",
        )

    def test_ability_categories(self):
        self.assertEqual(ability_category("文本定位", "意图清晰"), "文本-清晰")
        self.assertEqual(ability_category("文本定位", "意图模糊"), "文本-模糊")
        self.assertEqual(ability_category("图标定位", "意图清晰"), "图标-清晰")
        self.assertEqual(ability_category("图标定位", "意图模糊"), "图标-模糊")
        self.assertEqual(ability_category("拒答", ""), "拒答")

    def test_vla_modes(self):
        self.assertFalse(_uses_vla_combo("vla-basic"))
        self.assertTrue(_uses_vla_combo("vla-combo"))

    def _make_workbook(self, root, instructions):
        image_dir = root / "device_captures"
        image_dir.mkdir()
        Image.new("RGB", (201, 101), "white").save(image_dir / "CASE.png")
        source = root / "test.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = TASK_SHEET
        sheet.append(["任务指令", "图片ID", "UI-TREE", "二级能力", "三级能力", "结果输出"])
        for instruction in instructions:
            sheet.append([
                instruction,
                "CASE.png",
                "",
                "文本定位",
                "意图清晰",
                json.dumps({"action": "click", "bbox": [[90, 40, 110, 60]]}),
            ])
        workbook.save(source)
        return source

    def test_path_and_inline_xml_resolution(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            image_dir = root / "device_captures"
            image_dir.mkdir()
            image = image_dir / "CASE.png"
            Image.new("RGB", (20, 10), "white").save(image)
            self.assertEqual(
                resolve_image_path("CASE.png", workbook_dir=root, project_root=root),
                image.resolve(),
            )
            _, xml_root = load_ui_tree(
                '<hierarchy><node text="设置"/></hierarchy>',
                workbook_dir=root,
                project_root=root,
            )
            self.assertEqual(xml_root.tag, "hierarchy")

    def test_image_id_with_extension_resolves_from_device_captures(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            capture_dir = root / "device_captures"
            capture_dir.mkdir()
            image = capture_dir / "IQY_001.png"
            Image.new("RGB", (20, 10), "white").save(image)

            self.assertEqual(
                resolve_image_path(
                    "IQY_001.png",
                    workbook_dir=root,
                    project_root=root,
                ),
                image.resolve(),
            )

    def test_xml_name_resolves_from_device_captures(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            capture_dir = root / "device_captures"
            capture_dir.mkdir()
            xml_path = capture_dir / "IQY_001_0.xml"
            xml_path.write_text(
                '<hierarchy><node package="com.qiyi.video.pad"/></hierarchy>',
                encoding="utf-8",
            )

            resolved_path, xml_root = load_ui_tree(
                "IQY_001_0.xml",
                workbook_dir=root,
                project_root=root,
            )

            self.assertEqual(resolved_path, xml_path.resolve())
            self.assertEqual(xml_root.tag, "hierarchy")

    def test_evaluate_workbook_preserves_tasks_and_writes_results(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            image_dir = root / "device_captures"
            image_dir.mkdir()
            Image.new("RGB", (201, 101), "white").save(image_dir / "CASE.png")
            xml_path = image_dir / "CASE_0.xml"
            xml_path.write_text(
                '<hierarchy><node package="com.qiyi.video.pad"/></hierarchy>',
                encoding="utf-8",
            )
            source = root / "test.xlsx"
            output = root / "result.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = TASK_SHEET
            sheet.append(["任务指令", "图片ID", "UI-TREE", "二级能力", "三级能力", "结果输出"])
            sheet.append([
                "点击设置",
                "CASE.png",
                "CASE_0.xml",
                "图标定位",
                "意图清晰",
                json.dumps({"action": "click", "bbox": [[90, 40, 110, 60]]}, ensure_ascii=False),
            ])
            workbook.save(source)
            pipelines = {
                "xml": StaticPipeline(no_match_outcome("xml")),
                "ocr": StaticPipeline(click_outcome(pixel_x=100, pixel_y=50, source="ocr")),
                "vla": StaticPipeline(click_outcome(pixel_x=150, pixel_y=50, source="vla")),
            }
            records = evaluate_workbook(
                source,
                output_path=output,
                pipeline=pipelines,
                project_root=root,
            )
            self.assertEqual(len(records), 1)
            self.assertTrue(records[0].correct)
            self.assertEqual(pipelines["xml"].inputs[0].app_package, None)
            self.assertEqual(pipelines["ocr"].inputs[0].app_package, None)
            self.assertEqual(pipelines["vla"].inputs, [])
            self.assertEqual(records[0].selected_engine, "ocr")
            self.assertEqual(records[0].engine_correct, {"xml": False, "ocr": True})
            result = load_workbook(output, data_only=False)
            self.assertIn(TASK_SHEET, result.sheetnames)
            self.assertIn(DETAIL_SHEET, result.sheetnames)
            self.assertIn(SUMMARY_SHEET, result.sheetnames)
            self.assertEqual(result[DETAIL_SHEET]["G2"].value, "图标-清晰")
            self.assertEqual(result[DETAIL_SHEET]["M2"].value, True)
            self.assertEqual(result[DETAIL_SHEET]["T2"].value, False)
            self.assertEqual(result[DETAIL_SHEET]["U2"].value, True)
            self.assertIsNone(result[DETAIL_SHEET]["V2"].value)
            self.assertTrue(str(result[SUMMARY_SHEET]["B9"].value).startswith("="))
            self.assertEqual(
                [result[SUMMARY_SHEET].cell(row=row, column=1).value for row in range(18, 24)],
                ["文本-清晰", "文本-模糊", "图标-清晰", "图标-模糊", "拒答", "总体"],
            )
            self.assertIn("'评测明细'!M2:M2", result[SUMMARY_SHEET]["C20"].value)
            self.assertIn("'评测明细'!U2:U2", result[SUMMARY_SHEET]["G20"].value)

    def test_selected_wrong_action_stops_fallback_chain(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._make_workbook(root, ["点击设置"])
            pipelines = {
                "xml": StaticPipeline(click_outcome(pixel_x=150, pixel_y=50, source="xml")),
                "ocr": StaticPipeline(click_outcome(pixel_x=100, pixel_y=50, source="ocr")),
                "vla": StaticPipeline(click_outcome(pixel_x=100, pixel_y=50, source="vla")),
            }

            records = evaluate_workbook(
                source,
                output_path=root / "result.xlsx",
                pipeline=pipelines,
                project_root=root,
            )

            self.assertFalse(records[0].correct)
            self.assertEqual(records[0].selected_engine, "xml")
            self.assertEqual(pipelines["ocr"].inputs, [])
            self.assertEqual(pipelines["vla"].inputs, [])

    def test_parallel_strategy_runs_all_engines_concurrently_and_unions_results(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._make_workbook(root, ["点击设置"])
            tracker = EngineOverlapTracker()
            pipelines = {
                "xml": DelayedEnginePipeline(
                    click_outcome(pixel_x=150, pixel_y=50, source="xml"), tracker
                ),
                "ocr": DelayedEnginePipeline(
                    click_outcome(pixel_x=100, pixel_y=50, source="ocr"), tracker
                ),
                "vla": DelayedEnginePipeline(no_match_outcome("vla"), tracker),
            }
            output = root / "result.xlsx"

            records = evaluate_workbook(
                source,
                output_path=output,
                pipeline=pipelines,
                project_root=root,
                workers=1,
                engine_strategy="parallel",
            )

            record = records[0]
            self.assertGreater(tracker.peak_active, 1)
            self.assertEqual(len(pipelines["xml"].inputs), 1)
            self.assertEqual(len(pipelines["ocr"].inputs), 1)
            self.assertEqual(len(pipelines["vla"].inputs), 1)
            self.assertTrue(record.correct)
            self.assertEqual(record.selected_engine, "xml,ocr")
            self.assertEqual(
                record.engine_correct,
                {"xml": False, "ocr": True, "vla": False},
            )
            self.assertEqual(list(json.loads(record.actual_json)), ["xml", "ocr", "vla"])
            self.assertEqual(
                list(json.loads(record.pixel_action_json)), ["xml", "ocr", "vla"]
            )
            report = load_workbook(output, data_only=False)
            self.assertEqual(report[SUMMARY_SHEET]["E2"].value, "parallel")

    def test_parallel_strategy_keeps_engine_error_but_passes_on_other_result(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._make_workbook(root, ["解析失败"])
            pipelines = three_engine_pipelines(
                xml=ParseFailurePipeline(),
                ocr=StaticPipeline(
                    click_outcome(pixel_x=100, pixel_y=50, source="ocr")
                ),
            )

            records = evaluate_workbook(
                source,
                output_path=root / "result.xlsx",
                pipeline=pipelines,
                project_root=root,
                engine_strategy="parallel",
            )

            self.assertTrue(records[0].correct)
            self.assertEqual(records[0].selected_engine, "ocr")
            self.assertEqual(records[0].error, "")
            details = json.loads(records[0].engine_details_json)
            self.assertIn("model action is missing", details["xml"]["error"])

    def test_parallel_combo_runs_all_engines_and_injects_app_package(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            capture_dir = root / "device_captures"
            capture_dir.mkdir()
            Image.new("RGB", (201, 101), "white").save(capture_dir / "CASE.png")
            (capture_dir / "CASE.xml").write_text(
                '<hierarchy><node package="com.qiyi.video.pad"/></hierarchy>',
                encoding="utf-8",
            )
            source = root / "test.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = TASK_SHEET
            sheet.append(
                ["任务指令", "图片ID", "UI-TREE", "二级能力", "三级能力", "结果输出"]
            )
            sheet.append(
                [
                    "点击设置",
                    "CASE.png",
                    "CASE.xml",
                    "文本定位",
                    "意图清晰",
                    json.dumps({"action": "click", "bbox": [[90, 40, 110, 60]]}),
                ]
            )
            workbook.save(source)
            pipelines = {
                "xml": StaticPipeline(no_match_outcome("xml")),
                "ocr": StaticPipeline(no_match_outcome("ocr")),
                "vla": StaticPipeline(
                    click_outcome(pixel_x=100, pixel_y=50, source="vla")
                ),
            }

            output = root / "result.xlsx"
            records = evaluate_workbook(
                source,
                output_path=output,
                pipeline=pipelines,
                project_root=root,
                vla_mode="vla-combo",
                engine_strategy="parallel",
            )

            self.assertTrue(records[0].correct)
            self.assertEqual(records[0].engine_correct, {"xml": False, "ocr": False, "vla": True})
            self.assertEqual(pipelines["xml"].inputs[0].app_package, None)
            self.assertEqual(pipelines["ocr"].inputs[0].app_package, None)
            self.assertEqual(
                pipelines["vla"].inputs[0].app_package, "com.qiyi.video.pad"
            )
            report = load_workbook(output, data_only=False)
            summary = report[SUMMARY_SHEET]
            self.assertEqual(summary["B2"].value, "XML+OCR+VLA")
            self.assertEqual(summary["B3"].value, "vla-combo")
            self.assertEqual(summary["K17"].value, "XML+OCR并集正确数")
            self.assertEqual(summary["L17"].value, "XML+OCR并集成功率")
            self.assertIn("'评测明细'!T2:T2", summary["K18"].value)
            self.assertIn("'评测明细'!U2:U2", summary["K18"].value)
            self.assertIn("-COUNTIFS", summary["K18"].value)
            self.assertEqual(summary["J18"].value, "=IFERROR(I18/B18,0)")
            self.assertEqual(summary["L18"].value, "=IFERROR(K18/B18,0)")

    def test_vla_basic_does_not_inject_app_package(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            capture_dir = root / "device_captures"
            capture_dir.mkdir()
            Image.new("RGB", (201, 101), "white").save(capture_dir / "CASE.png")
            (capture_dir / "CASE.xml").write_text(
                '<hierarchy><node package="com.qiyi.video.pad"/></hierarchy>',
                encoding="utf-8",
            )
            source = root / "test.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = TASK_SHEET
            sheet.append(
                ["任务指令", "图片ID", "UI-TREE", "二级能力", "三级能力", "结果输出"]
            )
            sheet.append(
                [
                    "点击设置",
                    "CASE.png",
                    "CASE.xml",
                    "文本定位",
                    "意图清晰",
                    json.dumps({"action": "click", "bbox": [[90, 40, 110, 60]]}),
                ]
            )
            workbook.save(source)
            pipelines = three_engine_pipelines(
                vla=StaticPipeline(
                    click_outcome(pixel_x=100, pixel_y=50, source="vla")
                )
            )

            output = root / "result.xlsx"
            records = evaluate_workbook(
                source,
                output_path=output,
                pipeline=pipelines,
                project_root=root,
                vla_mode="vla-basic",
                engine_strategy="parallel",
            )

            self.assertTrue(records[0].correct)
            self.assertIsNone(pipelines["vla"].inputs[0].app_package)
            summary = load_workbook(output, data_only=False)[SUMMARY_SHEET]
            self.assertEqual(summary["B3"].value, "vla-basic")
            self.assertEqual(summary["K17"].value, "XML+OCR并集正确数")

    def test_missing_columns_are_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "bad.xlsx"
            workbook = Workbook()
            workbook.active.title = TASK_SHEET
            workbook.active.append(["任务指令"])
            workbook.save(source)
            with self.assertRaisesRegex(ValueError, "Missing task columns"):
                evaluate_workbook(
                    source,
                    output_path=root / "out.xlsx",
                    pipeline=three_engine_pipelines(),
                    project_root=root,
                )

    def test_concurrent_evaluation_overlaps_and_preserves_row_order(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._make_workbook(root, ["慢", "中", "快"])
            output = root / "result.xlsx"
            pipeline = ConcurrentPipeline(click_outcome(pixel_x=100, pixel_y=50))
            pipelines = three_engine_pipelines(vla=pipeline)

            records = evaluate_workbook(
                source,
                output_path=output,
                pipeline=pipelines,
                project_root=root,
                vla_mode="vla-basic",
                workers=3,
            )

            self.assertGreater(pipeline.peak_active, 1)
            self.assertEqual(pipeline.completion_order, ["快", "中", "慢"])
            self.assertEqual([record.source_row for record in records], [2, 3, 4])
            result = load_workbook(output, data_only=False)
            self.assertEqual(
                [result[DETAIL_SHEET].cell(row=row, column=1).value for row in range(2, 5)],
                [2, 3, 4],
            )

    def test_concurrent_failure_is_isolated(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._make_workbook(root, ["快", "失败", "中"])
            pipeline = ConcurrentPipeline(click_outcome(pixel_x=100, pixel_y=50))
            pipelines = three_engine_pipelines(vla=pipeline)

            records = evaluate_workbook(
                source,
                output_path=root / "result.xlsx",
                pipeline=pipelines,
                project_root=root,
                vla_mode="vla-basic",
                workers=3,
            )

            self.assertTrue(records[0].correct)
            self.assertEqual(records[1].error, "RuntimeError: planned failure")
            self.assertFalse(records[1].correct)
            self.assertTrue(records[2].correct)

    def test_unexpected_model_parse_failure_does_not_stop_batch(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._make_workbook(root, ["快", "解析失败", "中"])
            pipelines = three_engine_pipelines(vla=ParseFailurePipeline())

            records = evaluate_workbook(
                source,
                output_path=root / "result.xlsx",
                pipeline=pipelines,
                project_root=root,
                vla_mode="vla-basic",
                workers=3,
            )

            self.assertEqual(len(records), 3)
            self.assertTrue(records[0].correct)
            self.assertIn("KeyError", records[1].error)
            self.assertIn("model action is missing", records[1].error)
            self.assertFalse(records[1].correct)
            self.assertTrue(records[2].correct)

    def test_workers_must_be_positive(self):
        with self.assertRaises(SystemExit):
            parse_arguments(["test.xlsx", "--workers", "0"])
        with self.assertRaises(SystemExit):
            parse_arguments(["test.xlsx", "--workers", "-2"])
        defaults = parse_arguments(["test.xlsx"])
        self.assertEqual(defaults.workers, 1)
        self.assertEqual(defaults.vla_mode, "vla-combo")
        self.assertFalse(hasattr(defaults, "mode"))
        self.assertEqual(defaults.engine_strategy, "serial")
        self.assertEqual(
            parse_arguments(
                ["test.xlsx", "--engine-strategy", "parallel"]
            ).engine_strategy,
            "parallel",
        )
        with self.assertRaises(SystemExit):
            parse_arguments(["test.xlsx", "--engine-strategy", "invalid"])
        for vla_mode in ("vla-basic", "vla-combo"):
            self.assertEqual(
                parse_arguments(["test.xlsx", "--vla-mode", vla_mode]).vla_mode,
                vla_mode,
            )
        with self.assertRaises(SystemExit):
            parse_arguments(["test.xlsx", "--vla-mode", "invalid"])
        with self.assertRaises(SystemExit):
            parse_arguments(["test.xlsx", "--mode", "xml-ocr-vla"])
        with self.assertRaisesRegex(ValueError, "at least 1"):
            evaluate_workbook(
                Path("missing.xlsx"),
                output_path=Path("unused.xlsx"),
                pipeline=three_engine_pipelines(),
                project_root=Path.cwd(),
                workers=0,
            )
        with self.assertRaisesRegex(ValueError, "engine_strategy"):
            evaluate_workbook(
                Path("missing.xlsx"),
                output_path=Path("unused.xlsx"),
                pipeline=three_engine_pipelines(),
                project_root=Path.cwd(),
                engine_strategy="invalid",
            )
        with self.assertRaisesRegex(ValueError, "vla_mode"):
            evaluate_workbook(
                Path("missing.xlsx"),
                output_path=Path("unused.xlsx"),
                pipeline=three_engine_pipelines(),
                project_root=Path.cwd(),
                vla_mode="invalid",
            )

    def test_all_three_engine_pipelines_are_required(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = self._make_workbook(root, ["点击设置"])
            with self.assertRaisesRegex(
                ValueError, "Missing evaluator pipelines: vla"
            ):
                evaluate_workbook(
                    source,
                    output_path=root / "result.xlsx",
                    pipeline={
                        "xml": StaticPipeline(no_match_outcome("xml")),
                        "ocr": StaticPipeline(no_match_outcome("ocr")),
                    },
                    project_root=root,
                )


if __name__ == "__main__":
    unittest.main()
